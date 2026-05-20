"""
Config Builder - Flask backend
Interfaz visual drag & drop para configurar config_gastos.json
"""
import os, re, json, warnings, subprocess, shutil, sys, unicodedata
from datetime import date
from flask import Flask, render_template, jsonify, request, send_file
import db as _db

warnings.filterwarnings('ignore')

# BASE: directorio del script en desarrollo local.
# En Railway, apunta al volumen persistente via DATA_DIR para que los archivos
# (INFORME REAL, carpetas mensuales, configs) sobrevivan reinicios y redeploys.
_APP_DIR = os.path.dirname(os.path.abspath(__file__))
BASE     = os.environ.get('DATA_DIR', _APP_DIR)

# Al arrancar en Railway, copiar la plantilla al volumen si aÃƒÂºn no existe allÃƒÂ­
_TEMPLATE = os.path.join(_APP_DIR, 'INFORME REAL_2026 - FORMATO VERSION ORIGINAL.xlsx')
_INFORME_EN_BASE = os.path.join(BASE, 'INFORME REAL_2026 - FORMATO VERSION ORIGINAL.xlsx')
if BASE != _APP_DIR and os.path.exists(_TEMPLATE) and not os.path.exists(_INFORME_EN_BASE):
    os.makedirs(BASE, exist_ok=True)
    shutil.copy2(_TEMPLATE, _INFORME_EN_BASE)

CONFIG_PATH  = os.path.join(BASE, 'config_gastos.json')
INFORME_PATH = os.path.join(BASE, 'INFORME REAL_2026 - FORMATO VERSION ORIGINAL.xlsx')
SCRIPT_PATH  = os.path.join(_APP_DIR, 'procesar_todo.py')

# Copiar config al volumen si no existe allÃƒÂ­
for _cfg in ('config_gastos.json', 'config_5105.json'):
    _src = os.path.join(_APP_DIR, _cfg)
    _dst = os.path.join(BASE, _cfg)
    if BASE != _APP_DIR and os.path.exists(_src) and not os.path.exists(_dst):
        shutil.copy2(_src, _dst)

# Inicializar DB y migrar desde JSON si es la primera vez
_db.init_db()
_db.migrate_from_json()

MES_NOMBRES = ['ENERO','FEBRERO','MARZO','ABRIL','MAYO','JUNIO',
               'JULIO','AGOSTO','SEPTIEMBRE','OCTUBRE','NOVIEMBRE','DICIEMBRE']

app = Flask(__name__)


# ---------------------------------------------------------------------------
# Utilidades XLS
# ---------------------------------------------------------------------------
def celda_str(raw):
    if isinstance(raw, float) and raw == int(raw):
        return str(int(raw))
    return str(raw).strip()


def norm_text(s):
    """Normaliza texto para búsquedas tolerantes a acentos y mojibake."""
    t = str(s or '').lower().replace('ã©', 'e')
    t = unicodedata.normalize('NFKD', t)
    return ''.join(ch for ch in t if not unicodedata.combining(ch))


def leer_todos_codigos(filepath):
    """
    Extrae filas con codigo y valor de archivos XLS fuente.
    - Archivos 5105: solo muestra la seccion 99 GENERAL (ignora secciones PAEBUGA, etc.)
    - Archivos ESTADO DE RESULTADOS: solo muestra la seccion 99 GENERAL
    - Otros archivos: lee todas las filas
    Parsea el formato combinado "codigo + espacios + descripcion" en col A de archivos 5105.
    """
    try:
        import xlrd, re as _re
        nombre = os.path.basename(filepath).upper()
        wb = xlrd.open_workbook(filepath, formatting_info=True)
        ws = wb.sheet_by_name('Hoja 1')

        es_5105      = nombre.startswith('5105_')
        es_er        = nombre.startswith('ESTADO DE RESULTADOS')
        es_7205      = nombre.startswith('7205_')
        es_7105      = nombre.startswith('7105_')
        es_seccionado = es_7205 or es_7105  # misma estructura de secciones por contrato

        # Detectar columna de valor (Debitos o Netos) y columna de Creditos
        col_val = 9
        col_credito = None
        modo_valor = 'debito'  # 'debito' | 'netos'
        _found_col = False
        for r in range(min(25, ws.nrows)):
            for c in range(ws.ncols):
                txt = norm_text(ws.cell_value(r, c))
                if 'debito' in txt:
                    col_val = c
                    modo_valor = 'debito'
                    _found_col = True
                elif 'netos' in txt and not _found_col:
                    col_val = c
                    modo_valor = 'netos'
                    _found_col = True
                if 'credito' in txt:
                    col_credito = c
            if _found_col:
                break
        col_letra = chr(65 + col_val)

        # Inicio: primera fila con valor != 0 (salta filas de encabezado del reporte)
        inicio = 0
        for r in range(ws.nrows):
            try:
                if float(ws.cell_value(r, col_val) or 0) != 0:
                    inicio = r
                    break
            except Exception:
                pass

        # Para archivos 5105: detectar filas que son cabecera de nueva seccion
        # (ej: "PAEBUGA08 CONTRATO UT BUGA 2026") para insertar separadores visuales
        secciones_5105 = set()
        if es_5105:
            en_5105 = False
            for r in range(ws.nrows):
                a_raw = str(ws.cell_value(r, 0))
                a     = a_raw.strip()
                if a == '5105':
                    en_5105 = True
                    continue
                if en_5105:
                    stripped = a_raw.lstrip(' ')
                    if stripped and not _re.match(r'\d', stripped):
                        secciones_5105.add(r)
                        en_5105 = False  # reiniciar para detectar el siguiente bloque

        # Para archivos 7205 y 7105: detectar secciones por contrato.
        # Cabeceras de secciÃ³n: filas sin espacio inicial en col A que no son un cÃ³digo
        # puro de dÃ­gitos (ej: "99 GENERAL", "PAECAL20 CONTR. CONSORCIO...").
        secciones_seccionado = set()
        if es_seccionado:
            for r in range(ws.nrows):
                a_raw = str(ws.cell_value(r, 0))
                a_str = a_raw.strip()
                if (a_raw and not a_raw[0].isspace() and a_str
                        and not _re.match(r'^\d{3,8}$', a_str)):
                    secciones_seccionado.add(r)

        # Para ESTADO DE RESULTADOS: detectar cabeceras de U.N. por contrato/unidad.
        # Regla: filas sin espacio inicial en col A, no numÃ©ricas, y sin texto genÃ©rico
        # de encabezado/metadata (U.N., Cuentas, etc.).
        secciones_er = set()
        if es_er:
            omitir_exactos = {
                'U.N.', 'CUENTAS', '001 PRINCIPAL', '99 GENERAL',
                'LIBRO:', 'DIFERENCIA POR CUENTAS DE RESULTADO:'
            }
            for r in range(ws.nrows):
                a_raw = str(ws.cell_value(r, 0))
                a_str = a_raw.strip()
                if not a_str:
                    continue
                if a_str.upper() in omitir_exactos:
                    continue
                # Evitar metadatos del encabezado del reporte
                if ('CORPORACION' in a_str.upper() or 'CIFRAS EN' in a_str.upper()):
                    continue
                # Cabecera de secciÃ³n: sin sangrÃ­a y no numÃ©rico puro
                if (a_raw and not a_raw[0].isspace()
                        and not _re.match(r'^\d+(\.0)?$', a_str)):
                    secciones_er.add(r)

        items        = []
        seccion_actual = None  # secciÃ³n activa (None = secciÃ³n inicial del archivo)

        for r in range(inicio, ws.nrows):
            # Insertar separador de secciÃ³n para archivos 7205/7105
            if es_seccionado and r in secciones_seccionado:
                a_sep = str(ws.cell_value(r, 0)).strip()
                seccion_actual = a_sep
                items.append({
                    'codigo'     : '',
                    'descripcion': a_sep,
                    'valor'      : 0,
                    'celda'      : '',
                    'negrita'    : False,
                    'indent'     : 0,
                    'separador'  : True,
                })
                continue

            # Insertar separador de secciÃ³n para ESTADO DE RESULTADOS
            if es_er and r in secciones_er:
                a_sep = str(ws.cell_value(r, 0)).strip()
                seccion_actual = a_sep
                items.append({
                    'codigo'     : '',
                    'descripcion': a_sep,
                    'valor'      : 0,
                    'debito'     : 0,
                    'credito'    : 0,
                    'op_jk'      : False,
                    'celda'      : '',
                    'negrita'    : False,
                    'indent'     : 0,
                    'separador'  : True,
                })
                continue

            # Insertar separador de seccion para archivos 5105 y saltar la fila como dato
            if es_5105 and r in secciones_5105:
                a_sep = str(ws.cell_value(r, 0)).strip()
                seccion_actual = a_sep  # actualizar secciÃ³n activa
                items.append({
                    'codigo'     : '',
                    'descripcion': a_sep,
                    'valor'      : 0,
                    'celda'      : '',
                    'negrita'    : False,
                    'indent'     : 0,
                    'separador'  : True,
                })
                continue  # la fila cabecera solo se muestra como separador, no como dato

            raw_a     = ws.cell_value(r, 0)
            raw_a_str = str(raw_a)
            a_str     = raw_a_str.strip()

            if not a_str or a_str in ('None', '0.0', '0'):
                continue
            if isinstance(raw_a, float) and raw_a == int(raw_a):
                a_str = str(int(raw_a))

            # Valor neto por fila: Debitos (J) - Creditos (K)
            debito = 0.0
            credito = 0.0
            if ws.ncols > col_val:
                raw_j = ws.cell_value(r, col_val)
                try:
                    debito = float(raw_j) if raw_j else 0.0
                except (ValueError, TypeError):
                    debito = 0.0
            if col_credito is not None and ws.ncols > col_credito:
                raw_k = ws.cell_value(r, col_credito)
                try:
                    credito = float(raw_k) if raw_k else 0.0
                except (ValueError, TypeError):
                    credito = 0.0
            usa_resta_jk = (modo_valor == 'debito' and col_credito is not None)
            val = (debito - credito) if usa_resta_jk else debito

            # Codigo y descripcion
            # Archivos 5105: col A puede ser " 010102           GERENCIA" (todo junto)
            codigo = a_str
            desc   = ''
            if es_5105:
                m = _re.match(r'^(\d{3,8})\s{2,}(.+)', a_str)
                if m:
                    # Fila detalle: codigo y descripcion separados por 2+ espacios
                    codigo = m.group(1)
                    desc   = m.group(2).strip()
                else:
                    # Fila total (ej: "5105") o encabezado (ej: "99 GENERAL")
                    if ws.ncols > 1:
                        raw_b = ws.cell_value(r, 1)
                        desc  = str(raw_b).strip() if raw_b else ''
                    if not desc:
                        # "001 PRINCIPAL" o "99 GENERAL" viene todo en col A
                        parts = a_str.split(' ', 1)
                        if len(parts) == 2 and not parts[0].isdigit() is False:
                            codigo = parts[0]
                            desc   = parts[1]
            else:
                if ws.ncols > 1:
                    raw_b = ws.cell_value(r, 1)
                    desc  = str(raw_b).strip() if raw_b else ''

            # Indentacion: espacios iniciales en col A o nivel de sangria XL
            leading = len(raw_a_str) - len(raw_a_str.lstrip(' '))
            indent  = 0
            try:
                xf_ind = wb.xf_list[ws.cell_xf_index(r, 0)]
                indent = int(xf_ind.alignment.indent_level or 0)
            except Exception:
                pass
            if indent == 0 and leading > 0:
                indent = 1

            # Negrita en col A
            negrita = False
            try:
                xf_idx  = ws.cell_xf_index(r, 0)
                xf      = wb.xf_list[xf_idx]
                font    = wb.font_list[xf.font_index]
                negrita = bool(font.bold)
            except Exception:
                pass

            if codigo and (desc or val != 0.0):
                items.append({
                    'codigo'     : codigo,
                    'descripcion': desc[:70],
                    'valor'      : val,
                    'debito'     : debito,
                    'credito'    : credito,
                    'op_jk'      : bool(usa_resta_jk and abs(credito) > 0),

                    'celda'      : f'{col_letra}{r + 1}',
                    'negrita'    : negrita,
                    'indent'     : indent,
                    'seccion'    : seccion_actual,
                })

        return items
    except Exception as e:
        return [{'error': str(e)}]


# ---------------------------------------------------------------------------
# Endpoints
# ---------------------------------------------------------------------------
@app.route('/')
def index():
    return render_template('index.html')


@app.route('/api/sheets')
def api_sheets():
    """Retorna las hojas del INFORME REAL con sus filas (col A + col B)."""
    try:
        import openpyxl
        wb    = openpyxl.load_workbook(INFORME_PATH, data_only=True)
        config = json.load(open(CONFIG_PATH, encoding='utf-8'))
        config_map = {item['codigo_a']: item['sources'] for item in config['items']}

        # Cargar tambiÃƒÂƒÃ‚Â©n el Excel con fÃƒÂƒÃ‚Â³rmulas para detectar SUMs internos
        wb_f = openpyxl.load_workbook(INFORME_PATH, data_only=False)

        sheets = []
        HOJAS_EXCLUIR  = {'%DIST C', '%DIST C '}
        # Hojas con 3 cols/mes: solo cols 3,6,9,12... son columnas de valor
        # Las intermedias (4,5,7,8...) son ratios/presupuesto y no deben detectarse como fÃƒÂƒÃ‚Â³rmula
        HOJAS_3COL = {
            'CASINO', 'CALI', 'YUMBO', 'BUGA',
            'COMEDORES CALI', 'COMEDORES PALMIRA', 'COMEDORES VALLE',
            'CTAS EN PPACION', 'PYG TOTAL', 'RECREARTE'
        }

        # Filas a excluir en TODAS las hojas
        FILAS_EXCLUIR_GLOBAL = {
            'CORPORACI\u00d3N HACIA UN VALLE SOLIDARIO',
            'CORPORACI\u00d3N HACIA UN VALLE SOLIDARIO ',
            'DIAS DE ATENCI\u00d3N',
            'DIAS DE ATENCI\u00d3N ',
        }

        # Filas a excluir por hoja (col B) - titulos sin valor
        FILAS_EXCLUIR = {
            'GASTOS OPERATIVOS': {
                'CUOTA DE APOYO Y SOSTENIMIENTO',  # tÃƒÂƒÃ‚Â­tulo sin valor configurable
                'GASTOS DE PERSONAL',               # fÃƒÂƒÃ‚Â³rmula interna: suma filas 6-10
                'DIVERSOS',                         # fÃƒÂƒÃ‚Â³rmula interna: suma filas 13-16
                'TOTAL GASTOS',                     # fÃƒÂƒÃ‚Â³rmula interna: suma totales
            },
            'CASINO': {
                'Intereses', 'Descuentos Comerciales', 'otros', 'Flete',
                'Industrializados PROPIOS', 'Flete Industrializados TERCEROS',
                'Transporte Materia Prima - Cavasa', 'Flete Preparados PROPIOS',
                'Flete Preparados TERCEROS', 'Material de empaque'
            },
            'YUMBO': {
                # Totales y fÃƒÂƒÃ‚Â³rmulas internas
                'TOTAL RACIONES', 'INGRESOS BRUTOS', 'OPERACIONALES',
                'INDUSTRIAS MANUFACTURERAS', 'NO OPERACIONALES', 'FINANCIEROS',
                'DESCUENTOS - GASTO OPERACIONAL', 'OTROS', 'TOTAL INGRESOS NETOS',
                'COSTOS DE PRODUCCION', 'INSUMOS CI', 'COSTOS PREPARADOS',
                'INSUMOS CP', 'FLETES CP', 'OTROS COSTOS Y GASTOS INDIRECTOS',
                'COSTOS DE PERSONAL', 'HONORARIOS', 'SERVICIOS', 'GASTOS LEGALES',
                'ADECUACIONES E INSTALACIONES', 'DIVERSOS', 'UTILIDAD BRUTA',
                'GASTOS ADMINISTRATIVOS Y NO OPERACI', 'TOTAL AJUSTES', 'UTILIDAD NETA',
                'MANO DE OBRA DIRECTA',
                # Descuentos automÃƒÂƒÃ‚Â¡ticos por porcentaje
                'Procultura 0,5%', 'Prohospitales 1%', 'Rete Ica 0,6%',
                'Adulto Mayor 2%', 'Prodeporte 2,5%', '2% Juzgado de Familia',
                '0,5% Propacifico',
                # Cross-sheet: se muestran como es_interno en la UI (tarjeta morada)
                'PREPARADOS PROPIOS', 'Costos Indirectos de Personal',
                # Sin valor / no configurables
                'Retefuente', 'COSTOS INDUSTRIALIZADOS',
                'RACION PARA PREPARAR EN CASA (Ins)', 'REFRIGERIOS (Ins)',
                'FLETES CI', 'INDUSTRIALIZADOS PROPIOS', 'INDUSTRIALIZADOS TERCEROS',
            },
            'CALI': {
                'TOTAL RACIONES', 'INGRESOS BRUTOS', 'OPERACIONALES',
                'INDUSTRIAS MANUFACTURERAS', 'NO OPERACIONALES', 'FINANCIEROS',
                'DESCUENTOS - GASTO OPERACIONAL', 'OTROS', 'TOTAL INGRESOS NETOS',
                'COSTOS DE PRODUCCION', 'COSTOS INDUSTRIALIZADOS', 'INSUMOS CI',
                'FLETES CI', 'COSTOS PREPARADOS', 'INSUMOS CP', 'FLETES CP',
                'OTROS COSTOS Y GASTOS INDIRECTOS', 'COSTOS DE PERSONAL',
                'HONORARIOS', 'SERVICIOS', 'GASTOS LEGALES',
                'ADECUACIONES E INSTALACIONES', 'DIVERSOS',
                'UTILIDAD BRUTA', 'GASTOS ADMINISTRATIVOS Y NO OPERACI',
                'TOTAL AJUSTES', 'UTILIDAD NETA',
                'MANO DE OBRA DIRECTA',
                'PREPARADOS PROPIOS',               # calculado: GASTOS VEHICULOS * %DIST
                # Cross-sheet: se calculan automÃƒÂƒÃ‚Â¡ticamente desde otras hojas del libro
                'Costos Indirectos de Personal',
                # Cross-sheet: se muestran como es_interno en la UI (tarjeta morada)
                '2% Estampillas Prounivalle',
                '0,88% Rte Ica',
                '1% Estampilla Prohospital',
                '0,5% Propacifico',
                'Retefuente',
                '2% Juzgado de Familia',
            },
            'BUGA': {
                # Totales y fÃƒÂƒÃ‚Â³rmulas internas
                'TOTAL RACIONES', 'INGRESOS BRUTOS', 'OPERACIONALES',
                'INDUSTRIAS MANUFACTURERAS', 'NO OPERACIONALES', 'FINANCIEROS',
                'DESCUENTOS - GASTO OPERACIONAL', 'OTROS', 'TOTAL INGRESOS NETOS',
                'COSTOS DE PRODUCCION', 'COSTOS INDUSTRIALIZADOS', 'INSUMOS CI',
                'FLETES CI', 'COSTOS PREPARADOS', 'INSUMOS CP', 'FLETES CP',
                'OTROS COSTOS Y GASTOS INDIRECTOS', 'COSTOS DE PERSONAL',
                'HONORARIOS', 'ARRENDAMIENTOS', 'SEGUROS', 'SERVICIOS',
                'GASTOS LEGALES', 'MANTENIMIENTO Y REPARACIONES',
                'ADECUACIONES E INSTALACIONES', 'GASTOS DE VIAJE', 'DIVERSOS',
                'UTILIDAD BRUTA', 'GASTOS ADMINISTRATIVOS Y NO OPERACIONALES',
                'TOTAL AJUSTES', 'UTILIDAD NETA',
                'MANO DE OBRA DIRECTA',
                # Nota: filas (Fac) visibles para configuraciÃƒÂƒÃ‚Â³n manual en UI
                # Porcentajes auto-calculados sobre ingresos
                '1% Estampilla Prohospital', '2,5% Estampilla Pro Deporte',
                '1% Estampilla Pro Univalle', '3% Estampilla Adulto Mayor',
                '0,5% Estampilla Universidad del Pacifico',
                '0,66% Rete Ica', 'Retefuente', '2% Juzgado de Familia',
                # Cross-sheet: se muestran como es_interno en la UI (tarjeta morada)
                'PREPARADOS PROPIOS',
                # Sin valor / no configurables
                'Intereses', 'Descuentos Comerciales', 'otros Fin',
                'INDUSTRIALIZADOS TERCEROS',
            },
            'RECREARTE': {
                # Totales y subtotales (fÃƒÂƒÃ‚Â³rmulas internas)
                'TOTAL RACIONES', 'INGRESOS BRUTOS',
                'OPERACIONALES', 'INDUSTRIAS MANUFACTURERAS',
                'NO OPERACIONALES', 'FINANCIEROS',
                'DESCUENTOS - GASTO OPERACIONAL', 'OTROS', 'TOTAL INGRESOS NETOS',
                'COSTOS DIRECTOS', 'COSTOS INDUSTRIALIZADOS', 'INSUMOS',
                'COSTOS PREPARADOS',
                'GASTOS DE PERSONAL', 'OTROS COSTOS Y GASTOS INDIRECTOS',
                'COSTOS DE PERSONAL', 'HONORARIOS', 'ARRENDAMIENTOS',
                'SEGUROS', 'SERVICIOS', 'GASTOS LEGALES',
                'MANTENIMIENTO Y REPARACIONES', 'ADECUACIONES E INSTALACIONES',
                'GASTOS DE VIAJE', 'DIVERSOS', 'TOTAL GASTOS',
            }
        }

        # Filas que se muestran como tÃƒÂ­tulo de secciÃƒÂ³n (separador visual, no editables)
        TITULOS_SECCION = {
            'GASTOS ADMINISTRATIVOS': {
                'GASTOS DE PERSONAL', 'HONORARIOS', 'IMPUESTOS', 'ARRENDAMIENTOS',
                'SEGUROS', 'SERVICIOS', 'GASTOS LEGALES', 'MANTENIMIENTO Y REPARACIONES',
                'ADECUACIONES E INSTALACIONES', 'GASTOS DE VIAJE', 'DIVERSOS',
                'OTROS GASTOS DIVERSOS', 'FINANCIEROS',
                # CASINO y RECREARTE son referencias internas, NO tÃƒÂ­tulos de secciÃƒÂ³n
            },
            'GASTOS VEHICULOS': {
                'GASTOS DE PERSONAL', 'IMPUESTOS', 'SEGUROS', 'SERVICIOS',
                'MANTENIMIENTO Y REPARACIONES', 'DIVERSOS',
            },
            'CASINO': {
                'INGRESOS BRUTOS', 'OPERACIONALES', 'INDUSTRIAS MANUFACTURERAS',
                'NO OPERACIONALES', 'FINANCIEROS', 'DESCUENTOS - GASTO OPERACIONAL', 'OTROS',
                'COSTOS DE PRODUCCION', 'COSTOS ALMUERZO', 'INSUMOS',
                'MANO DE OBRA DIRECTA', 'OTROS COSTOS Y GASTOS INDIRECTOS',
                'COSTOS DE PERSONAL', 'HONORARIOS', 'ARRENDAMIENTOS', 'SEGUROS',
                'SERVICIOS', 'GASTOS LEGALES', 'MANTENIMIENTO Y REPARACIONES',
                'ADECUACIONES E INSTALACIONES', 'GASTOS DE VIAJE', 'DIVERSOS',
            },
            'CALI': {
                'ARRENDAMIENTOS', 'SEGUROS', 'MANTENIMIENTO Y REPARACIONES',
                'GASTOS DE VIAJE', 'GASTOS ADMINISTRATIVOS Y NO OPERACIONALES',
            },
            'YUMBO': {
                'ARRENDAMIENTOS', 'SEGUROS', 'MANTENIMIENTO Y REPARACIONES',
                'GASTOS DE VIAJE', 'GASTOS ADMINISTRATIVOS Y NO OPERACIONALES',
            },
            'COMEDORES CALI': {
                'INGRESOS BRUTOS', 'OPERACIONALES', 'INDUSTRIAS MANUFACTURERAS',
                'NO OPERACIONALES', 'FINANCIEROS', 'DESCUENTOS - GASTO OPERACIONAL', 'OTROS',
                'COSTOS DE PRODUCCION', 'COSTOS INDUSTRIALIZADOS', 'INSUMOS CI', 'FLETES OPERADORES',
                'COSTOS PREPARADOS', 'INSUMOS CP', 'FLETES',
                'MANO DE OBRA DIRECTA', 'OTROS COSTOS Y GASTOS INDIRECTOS',
                'COSTOS DE PERSONAL', 'HONORARIOS', 'ARRENDAMIENTOS', 'SEGUROS',
                'SERVICIOS', 'GASTOS LEGALES', 'MANTENIMIENTO Y REPARACIONES',
                'ADECUACIONES E INSTALACIONES', 'GASTOS DE VIAJE', 'DIVERSOS',
                'GASTOS ADMINISTRATIVOS Y NO OPERACIONALES',
            },
            'COMEDORES PALMIRA': {
                'INGRESOS BRUTOS', 'OPERACIONALES', 'INDUSTRIAS MANUFACTURERAS',
                'NO OPERACIONALES', 'FINANCIEROS', 'DESCUENTOS - GASTO OPERACIONAL', 'OTROS',
                'COSTOS DE PRODUCCION', 'COSTOS INDUSTRIALIZADOS', 'INSUMOS', 'FLETES',
                'COSTOS PREPARADOS',
                'MANO DE OBRA DIRECTA', 'OTROS COSTOS Y GASTOS INDIRECTOS',
                'COSTOS DE PERSONAL', 'HONORARIOS', 'ARRENDAMIENTOS', 'SEGUROS',
                'SERVICIOS', 'GASTOS LEGALES', 'MANTENIMIENTO Y REPARACIONES',
                'ADECUACIONES E INSTALACIONES', 'GASTOS DE VIAJE', 'DIVERSOS',
                'GASTOS ADMINISTRATIVOS Y NO OPERACIONALES',
            },
            'CTAS EN PPACION': {
                'INGRESOS BRUTOS', 'OPERACIONALES', 'INDUSTRIAS MANUFACTURERAS',
                'NO OPERACIONALES', 'FINANCIEROS', 'DESCUENTOS - GASTO OPERACIONAL', 'OTROS',
                'COSTOS DE PRODUCCION', 'COSTOS INDUSTRIALIZADOS', 'INSUMOS',
                'FLETES - OTROS OPERADORES', 'COSTOS PREPARADOS', 'FLETES - CHVS',
                'MANO DE OBRA DIRECTA', 'OTROS COSTOS Y GASTOS INDIRECTOS',
                'COSTOS DE PERSONAL', 'HONORARIOS', 'ARRENDAMIENTOS', 'SEGUROS',
                'SERVICIOS', 'GASTOS LEGALES', 'MANTENIMIENTO Y REPARACIONES',
                'ADECUACIONES E INSTALACIONES', 'GASTOS DE VIAJE', 'DIVERSOS',
                'GASTOS ADMINISTRATIVOS Y NO OPERACIONALES',
            },
            'PYG TOTAL': {
                'INGRESOS BRUTOS', 'OPERACIONALES', 'NO OPERACIONALES', 'FINANCIEROS',
                'DESCUENTOS - GASTO OPERACIONAL ADMON', 'OTROS',
                'COSTOS DE PRODUCCION', 'COSTOS INDUSTRIALIZADOS', 'INSUMOS INDUSTRIALIZADOS',
                'FLETES OPERADORES', 'COSTOS PREPARADOS', 'INSUMOS PREPARADOS', 'FLETES',
                'MANO DE OBRA DIRECTA', 'OTROS COSTOS Y GASTOS INDIRECTOS',
                'COSTOS DE PERSONAL', 'HONORARIOS', 'ARRENDAMIENTOS', 'SEGUROS',
                'SERVICIOS', 'GASTOS LEGALES', 'MANTENIMIENTO Y REPARACIONES',
                'ADECUACIONES E INSTALACIONES', 'GASTOS DE VIAJE', 'DIVERSOS',
                'GASTOS ADMINISTRATIVOS Y NO OPERACIONALES',
            },
        }

        for sh_name in wb.sheetnames:
            ws      = wb[sh_name]
            ws_f    = wb_f[sh_name]
            if ws.sheet_state == 'hidden':
                continue
            if sh_name.strip() in HOJAS_EXCLUIR:
                continue
            rows = []
            for row in ws.iter_rows():
                a = str(row[0].value or '').strip()
                b = str(row[1].value or '').strip() if len(row) > 1 else ''

                # Incluir si col A tiene cÃƒÂƒÃ‚Â³digo numÃƒÂƒÃ‚Â©rico O si col B tiene texto descriptivo
                tiene_codigo = bool(a and re.match(r'^\d', a))
                if not tiene_codigo and not b:
                    continue  # fila vacÃƒÂƒÃ‚Â­a

                # Excluir filas globales y especÃƒÂƒÃ‚Â­ficas por hoja
                if b in FILAS_EXCLUIR_GLOBAL:
                    continue
                hoja_limpia = sh_name.strip()
                if b in FILAS_EXCLUIR.get(hoja_limpia, set()):
                    continue

                # Detectar fÃƒÂƒÃ‚Â³rmula interna: referencia otras filas de la MISMA hoja
                # Solo se revisan las columnas de valor (no las de ratio/presupuesto)
                es_formula = False
                es_interno  = False
                current_row = row[0].row
                if hoja_limpia in HOJAS_3COL:
                    cols_valor = [3, 6, 9, 12, 15, 18, 21, 24, 27, 30, 33, 36]
                else:
                    cols_valor = list(range(3, 15))
                for col_num in cols_valor:
                    v = ws_f.cell(row=current_row, column=col_num).value
                    if not v or not isinstance(v, str) or not v.startswith('='):
                        continue
                    # Referencia a libro externo ([n]): ignorar completamente
                    if '[' in v:
                        continue
                    # Referencia interna sin comillas (ej: =CASINO!$C$144)
                    if '!' in v:
                        es_interno = True
                        continue
                    # Referencia a otras filas de la misma hoja: formula interna
                    nums = re.findall(r'\d+', v)
                    otras_filas = [int(n) for n in nums if int(n) != current_row and 1 < int(n) < 500]
                    if otras_filas:
                        es_formula = True
                        break

                # Titulos de seccion: formula interna pero se muestran como separador visual
                es_titulo = es_formula and b in TITULOS_SECCION.get(hoja_limpia, set())

                # Filas con formulas internas: Excel las calcula solo, no se muestran.
                if es_formula and not es_titulo:
                    continue

                # Clave de busqueda en config
                clave = re.sub(r'\s+', ' ', a) if tiene_codigo else f'B:{b}'
                fuentes = config_map.get(clave, config_map.get(a, []))

                rows.append({
                    'fila'        : row[0].row,
                    'codigo'      : a,
                    'descripcion' : b,
                    'fuentes'     : fuentes,
                    'tiene_codigo': tiene_codigo,
                    'es_titulo'   : es_titulo,
                    'es_interno'  : es_interno,
                })

            if rows:
                sheets.append({'nombre': sh_name.strip(), 'filas': rows})
        return jsonify(sheets)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/folders')
def api_folders():
    """Lista carpetas mensuales disponibles en BASE."""
    carpetas = []
    for nombre in sorted(os.listdir(BASE)):
        ruta = os.path.join(BASE, nombre)
        if not os.path.isdir(ruta):
            continue
        nombre_limpio = nombre.replace('-DL', '').upper()
        if nombre_limpio in MES_NOMBRES:
            # Contar archivos XLS
            n_xls = len([f for f in os.listdir(ruta) if f.lower().endswith('.xls')])
            if n_xls > 0:
                carpetas.append({'nombre': nombre, 'n_archivos': n_xls})
    return jsonify(carpetas)


@app.route('/api/files/<path:folder>')
def api_files(folder):
    """Lee todos los XLS de la carpeta y extrae cÃƒÂƒÃ‚Â³digos por archivo."""
    carpeta = os.path.join(BASE, folder) if not os.path.isabs(folder) else folder
    if not os.path.isdir(carpeta):
        return jsonify({'error': f'Carpeta no encontrada: {carpeta}'}), 404

    resultado = {}
    for nombre in sorted(os.listdir(carpeta)):
        if not nombre.lower().endswith('.xls'):
            continue
        ruta  = os.path.join(carpeta, nombre)
        items = leer_todos_codigos(ruta)
        resultado[nombre] = {'items': items}

    return jsonify(resultado)


@app.route('/api/config')
def api_config():
    return jsonify(_db.load_config())


@app.route('/api/config/save', methods=['POST'])
def api_config_save():
    """Recibe el nuevo config y lo guarda en la DB."""
    try:
        nuevo_config = request.get_json()
        _db.save_config(nuevo_config)
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/historial')
@app.route('/api/historial/<mes>')
def api_historial(mes=None):
    """Retorna el historial de ejecuciones, opcionalmente filtrado por mes."""
    try:
        return jsonify(_db.get_historial(mes))
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/run/<path:folder>', methods=['POST'])
def api_run(folder):
    """Ejecuta procesar_todo.py para la carpeta indicada."""
    try:
        env = os.environ.copy()
        env['CHVS_BASE'] = BASE
        result = subprocess.run(
            [sys.executable, SCRIPT_PATH, folder],
            capture_output=True, text=True, encoding='utf-8', errors='replace',
            timeout=300, env=env
        )
        return jsonify({
            'ok'    : result.returncode == 0,
            'stdout': result.stdout,
            'stderr': result.stderr
        })
    except subprocess.TimeoutExpired:
        return jsonify({'error': 'Timeout - el script tardo mas de 5 minutos'}), 500
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/upload/<folder>', methods=['POST'])
def api_upload(folder):
    """Sube archivos XLS a una carpeta mensual en el servidor."""
    folder_clean = folder.upper().replace('-DL', '').replace('-', '')
    if folder_clean not in MES_NOMBRES:
        return jsonify({'error': f'Nombre de carpeta invÃƒÂ¡lido: {folder}'}), 400
    carpeta = os.path.join(BASE, folder)
    os.makedirs(carpeta, exist_ok=True)
    # Limpiar archivos XLS previos para evitar mezclar con carga anterior
    for existing in os.listdir(carpeta):
        if existing.lower().endswith('.xls') or existing.lower().endswith('.xlsx'):
            os.remove(os.path.join(carpeta, existing))
    files = request.files.getlist('files')
    saved = []
    for f in files:
        nombre = f.filename
        if nombre.lower().endswith('.xls') or nombre.lower().endswith('.xlsx'):
            f.save(os.path.join(carpeta, nombre))
            saved.append(nombre)
    return jsonify({'ok': True, 'saved': saved, 'folder': folder, 'total': len(saved)})


@app.route('/api/upload/informe', methods=['POST'])
def api_upload_informe():
    """Reemplaza el INFORME REAL base en el volumen con el archivo subido."""
    f = request.files.get('file')
    if not f:
        return jsonify({'error': 'No se recibio ningun archivo'}), 400
    nombre = f.filename or ''
    if not nombre.lower().endswith('.xlsx'):
        return jsonify({'error': 'El archivo debe ser .xlsx'}), 400
    f.save(INFORME_PATH)
    return jsonify({'ok': True, 'mensaje': 'INFORME REAL reemplazado correctamente'})


@app.route('/api/download/informe')
def api_download_informe():
    """Descarga el INFORME REAL actualizado."""
    if not os.path.exists(INFORME_PATH):
        return jsonify({'error': 'Archivo no encontrado'}), 404
    return send_file(INFORME_PATH, as_attachment=True,
                     download_name='INFORME REAL_2026 - FORMATO VERSION ORIGINAL.xlsx')


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    debug = os.environ.get('RAILWAY_ENVIRONMENT') is None  # False en Railway
    print(f'\n  Config Builder iniciado en puerto {port}')
    if debug:
        print(f'  Abre http://localhost:{port} en tu navegador\n')
    app.run(debug=debug, host='0.0.0.0', port=port)




