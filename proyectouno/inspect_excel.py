import openpyxl
import os

file_path = r"c:\Users\User\OneDrive\Desktop\CHVS\adriana\librouno_ejemplo.xlsx"

try:
    # load_workbook with data_only=False to see formulas
    wb = openpyxl.load_workbook(file_path, data_only=False)
    print(f"Excel File: {os.path.basename(file_path)}")
    print(f"Sheets: {wb.sheetnames}")
    
    # Check for external links targets
    print("\n--- EXTERNAL LINKS TARGETS ---")
    if hasattr(wb, '_external_links'):
        for i, link in enumerate(wb._external_links):
            print(f"Link {i}: {link.file_link.Target}")
    else:
        print("No external links found in workbook metadata.")

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        print(f"\n--- Sheet: {sheet_name} ---")
        
        # Merged cells
        merged_ranges = list(sheet.merged_cells.ranges)
        print(f"Merged cells count: {len(merged_ranges)}")
        if merged_ranges:
            print(f"Example merged ranges: {', '.join([str(r) for r in merged_ranges[:10]])}")
            
        # Search for formulas
        formula_cells = []
        external_formula_cells = []
        
        # Iterating through all cells to find formulas
        # This can be slow for very large sheets, but should be fine for typical files.
        for row in sheet.iter_rows():
            for cell in row:
                val = cell.value
                if val and isinstance(val, str) and val.startswith('='):
                    formula_cells.append((cell.coordinate, val))
                    if '[' in val or '.xls' in val:
                        external_formula_cells.append((cell.coordinate, val))
                        
        print(f"Total formula cells: {len(formula_cells)}")
        print(f"External formula cells: {len(external_formula_cells)}")
        
        if external_formula_cells:
            print("External formula samples (first 10):")
            for coord, formula in external_formula_cells[:10]:
                print(f"  {coord}: {formula}")
                
except Exception as e:
    import traceback
    print(f"Error analyzing Excel: {e}")
    traceback.print_exc()
