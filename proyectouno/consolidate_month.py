import openpyxl
import xlrd
import os
import re
import sys

# CONFIGURACIÓN
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
MASTER_FILE = os.path.join(BASE_DIR, "librouno_ejemplo.xlsx")

# Mapeo de columnas por mes
MONTH_COLS = {
    'ENERO': 'B', 'FEBRERO': 'C', 'MARZO': 'D', 'ABRIL': 'E', 
    'MAYO': 'F', 'JUNIO': 'G', 'JULIO': 'H', 'AGOSTO': 'I', 
    'SEPTIEMBRE': 'J', 'OCTUBRE': 'K', 'NOVIEMBRE': 'L', 'DICIEMBRE': 'M'
}

def get_xls_value(folder_path, filename_pattern, sheet_name, cell_ref, found_files):
    """Extrae un valor de un archivo .xls antiguo usando patrones de nombre."""
    try:
        # Buscar el archivo que coincida con el patrón en la carpeta del mes
        matching_files = [f for f in os.listdir(folder_path) if f.lower().startswith(filename_pattern.lower())]
        if not matching_files:
            return 0
        
        filename = matching_files[0]
        found_files.add(filename)
        file_path = os.path.join(folder_path, filename)
        
        wb = xlrd.open_workbook(file_path)
        try:
            sheet = wb.sheet_by_name(sheet_name)
        except xlrd.biffh.XLRDError:
            # Si no encuentra "Hoja 1", intenta con la primera hoja disponible
            sheet = wb.sheet_by_index(0)
            
        # Convertir cell_ref (ej. $I$9) a índices (row, col)
        clean_ref = cell_ref.replace('$', '')
        col_letter = re.match(r"([A-Z]+)", clean_ref).group(1)
        row_num = int(re.search(r"(\d+)", clean_ref).group(1)) - 1
        
        col_idx = 0
        for char in col_letter:
            col_idx = col_idx * 26 + (ord(char) - ord('A') + 1)
        col_idx -= 1
        
        val = sheet.cell_value(row_num, col_idx)
        # Asegurarse de devolver un número, o 0 si está vacío
        if val == "" or val is None:
            return 0
        return float(val)
    except Exception:
        return 0

def automate_month(target_month_name):
    target_month_name = target_month_name.upper()
    print(f"\n>>> INICIANDO PROCESO PARA: {target_month_name} <<<")
    
    if not os.path.exists(MASTER_FILE):
        print(f"ERROR: No se encuentra el archivo maestro: {MASTER_FILE}")
        return

    wb_master = openpyxl.load_workbook(MASTER_FILE, data_only=False)
    sheet = wb_master['GASTOS ADMINISTRATIVOS ']
    
    dest_col_letter = MONTH_COLS.get(target_month_name)
    if not dest_col_letter:
        print(f"ERROR: '{target_month_name}' no es un mes válido.")
        return

    dest_col_idx = openpyxl.utils.column_index_from_string(dest_col_letter)
    folder_path = os.path.join(BASE_DIR, target_month_name)
    
    if not os.path.exists(folder_path):
        print(f"ERROR: No existe la carpeta para el mes: {folder_path}")
        return

    # Mapeo de patrones de archivos
    link_map = {}
    if hasattr(wb_master, '_external_links'):
        for i, link in enumerate(wb_master._external_links):
            target = os.path.basename(link.file_link.Target).replace('%20', ' ')
            parts = target.split('_')
            if len(parts) > 2:
                # Agregamos el guion bajo al final para forzar la coincidencia exacta
                # Evita que "UTALIM YUMBO" coincida con "UTALIM YUMBO 2025"
                pattern = parts[0] + "_" + parts[1] + "_"
            else:
                pattern = target.replace(".xls", "").replace(".xlsx", "")
            link_map[str(i+1)] = pattern

    found_files = set()
    total_rows_processed = 0
    total_cells_updated = 0

    print("Analizando y traduciendo fórmulas de la plantilla...")

    for row in range(4, sheet.max_row + 1):
        total_rows_processed += 1
        # Usamos siempre la columna C (Febrero) como plantilla
        cell_template = sheet.cell(row=row, column=3) 
        formula = str(cell_template.value) if cell_template.value else ""
        
        if formula.startswith('='):
            # 1. Traducir las referencias de celdas internas de la Columna C a la nueva columna
            # Ej: =C96+C109 -> =D96+D109 (si dest_col_letter es 'D')
            # Usamos una expresión regular para buscar letras 'C' seguidas de un número, asegurando que sea una palabra completa (\b)
            # Y que no esté dentro de comillas simples (para no dañar nombres de hojas).
            
            def replace_internal_ref(match):
                # match.group(0) es algo como 'C96'
                return f"{dest_col_letter}{match.group(1)}"
            
            # Reemplaza C seguido de números solo si no tiene $ alrededor, o si es una referencia relativa normal
            # \bC(\d+)\b busca exactamente "C" seguido de números
            translated_formula = re.sub(r"\bC(\d+)\b", replace_internal_ref, formula)
            
            # 2. Encontrar y evaluar todos los vínculos externos en la fórmula
            # Ej: '[4]Hoja 1'!$L$8
            def evaluate_external_link(match):
                link_id = match.group(1)
                sheet_n = match.group(2)
                cell_ref = match.group(3)
                
                file_pattern = link_map.get(link_id)
                if file_pattern:
                    val = get_xls_value(folder_path, file_pattern, sheet_n, cell_ref, found_files)
                    # Devuelve el valor como string para inyectarlo en la fórmula de Excel
                    return str(val)
                return "0"

            # Busca cosas como '[4]Hoja 1'!$L$8
            final_formula = re.sub(r"'?\[(\d+)\]([^']+)'?!([A-Z0-9\$]+)", evaluate_external_link, translated_formula)
            
            # Escribimos la fórmula resultante en la celda destino
            # Si era una fórmula matemática con externos, quedará como "=15000+25000-6457629"
            # Si era solo una referencia interna, quedará como "=D96+D109"
            sheet.cell(row=row, column=dest_col_idx).value = final_formula
            total_cells_updated += 1
            
        elif formula:
            # Si era un valor fijo (ej. "FEB", 318000), lo copiamos tal cual
            if isinstance(cell_template.value, (int, float)):
                sheet.cell(row=row, column=dest_col_idx).value = cell_template.value
            elif isinstance(cell_template.value, str):
                # Reemplazamos FEB por el nombre del mes si es el caso (Fila 4)
                if cell_template.value == "FEB":
                    sheet.cell(row=row, column=dest_col_idx).value = target_month_name[:3]
                else:
                    sheet.cell(row=row, column=dest_col_idx).value = cell_template.value
            total_cells_updated += 1

    # Guardar cambios
    output_name = f"librouno_consolidado_{target_month_name}_FINAL.xlsx"
    output_path = os.path.join(BASE_DIR, output_name)
    
    # Intento de guardado seguro
    try:
        wb_master.save(output_path)
    except PermissionError:
        output_name = f"librouno_consolidado_{target_month_name}_FINAL_V2.xlsx"
        output_path = os.path.join(BASE_DIR, output_name)
        wb_master.save(output_path)

    # REPORTE FINAL
    print("\n" + "="*40)
    print(f"RESUMEN DE PROCESAMIENTO: {target_month_name}")
    print("="*40)
    print(f"1. Archivos detectados en carpeta {target_month_name}:")
    all_files = [f for f in os.listdir(folder_path) if f.lower().endswith('.xls')]
    for f in all_files:
        status = "[OK]" if f in found_files else "[NO USADO]"
        print(f"   {status} {f}")
    
    print(f"\n2. Estadísticas:")
    print(f"   - Filas analizadas: {total_rows_processed}")
    print(f"   - Celdas generadas con fórmulas/valores: {total_cells_updated}")
    print(f"   - Archivo generado: {output_name}")
    print("="*40)
    print("\nPROCESO COMPLETADO CON ÉXITO.")

if __name__ == "__main__":
    if len(sys.argv) > 1:
        automate_month(sys.argv[1])
    else:
        mes = input("Ingrese el nombre del mes a procesar (ej: MARZO): ")
        automate_month(mes)
