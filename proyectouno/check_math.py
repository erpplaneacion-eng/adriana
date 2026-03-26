import openpyxl

file_path = r"c:\Users\User\OneDrive\Desktop\CHVS\adriana\librouno_ejemplo.xlsx"

wb = openpyxl.load_workbook(file_path, data_only=False)
sheet = wb['GASTOS ADMINISTRATIVOS ']

print("--- Fórmulas Complejas en Columna C ---")
for row in range(4, sheet.max_row + 1):
    cell = sheet.cell(row=row, column=3)
    formula = str(cell.value) if cell.value else ""
    
    if formula.startswith('='):
        # Count external links
        num_links = formula.count('.xls') + formula.count(']')
        # Check for math operators outside of string paths
        if num_links > 1 or '-' in formula or '+' in formula:
            print(f"Celda C{row} (Fila {row}): {formula}")

