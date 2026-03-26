import openpyxl
import os
import re

file_path = r"c:\Users\User\OneDrive\Desktop\CHVS\adriana\librouno_ejemplo.xlsx"

try:
    wb = openpyxl.load_workbook(file_path, data_only=False)
    sheet = wb['GASTOS ADMINISTRATIVOS ']
    
    # Map link numbers to filenames for better readability
    # Link targets are stored in wb._external_links
    link_map = {}
    if hasattr(wb, '_external_links'):
        for i, link in enumerate(wb._external_links):
            # Extract filename from target (handles URL-encoded paths)
            target = link.file_link.Target
            filename = os.path.basename(target).replace('%20', ' ')
            link_map[str(i+1)] = filename

    print(f"{'Fila':<5} | {'Descripción (Col A)':<40} | {'Contenido Col C (Fórmula / Archivo)':<60}")
    print("-" * 115)

    # We start from row 4 assuming 1-3 are headers based on previous analysis (A2:A3, B2:N3 merged)
    for row_idx in range(4, sheet.max_row + 1):
        cell_a = sheet.cell(row=row_idx, column=1)
        cell_c = sheet.cell(row=row_idx, column=3)
        
        desc = str(cell_a.value).strip() if cell_a.value else ""
        formula = str(cell_c.value) if cell_c.value else ""
        
        if not desc and not formula:
            continue

        # Identify external files in the formula
        # Formulas look like ='[4]Hoja 1'!$L$8
        external_refs = re.findall(r"\[(\d+)\]", formula)
        files_involved = [link_map.get(ref, f"Link {ref}") for ref in external_refs]
        
        if files_involved:
            relation = f"VINCULADO A: {', '.join(files_involved)}"
        elif formula.startswith('='):
            relation = "Fórmula Interna"
        elif formula:
            relation = f"Valor Fijo: {formula}"
        else:
            relation = "Vacío"

        # Show only relevant rows (those with description or formula)
        if desc or formula:
            print(f"{row_idx:<5} | {desc[:40]:<40} | {relation}")

except Exception as e:
    print(f"Error: {e}")
