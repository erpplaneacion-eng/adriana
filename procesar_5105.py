"""
procesar_5105.py
----------------
Lee los archivos 5105_*.xls de una carpeta mensual, aplica las reglas
de cada entidad (definidas en config_5105.json) y escribe el resultado
en la columna correspondiente del archivo principal de Gastos Administrativos.

Uso:
    py procesar_5105.py                     # usa FEBRERO-DL por defecto
    py procesar_5105.py MARZO-DL            # nombre de carpeta relativo
    py procesar_5105.py C:\\ruta\\MARZO-DL  # ruta absoluta
"""

import os
import re
import sys
import json
import warnings
import xlrd
import openpyxl

warnings.filterwarnings('ignore')

# ---------------------------------------------------------------------------
# Configuración de rutas
# ---------------------------------------------------------------------------
BASE = r'C:\Users\User\OneDrive\Desktop\CHVS\adriana'
ARCHIVO_PRINCIPAL = os.path.join(BASE, 'ANALIIS PESTAÑA GASTOS ADMINISTRATIVOS.xlsx')
CONFIG_PATH = os.path.join(BASE, 'config_5105.json')

# ---------------------------------------------------------------------------
# Mapeo mes → columna en el xlsx principal (openpyxl usa índice 1-based)
# C=ENE, D=FEB, E=MAR, F=ABR, G=MAY, H=JUN, I=JUL, J=AGO, K=SEP, L=OCT, M=NOV, N=DIC
# ---------------------------------------------------------------------------
MES_COLUMNA = {
    "ENERO": 3, "FEBRERO": 4, "MARZO": 5, "ABRIL": 6,
    "MAYO": 7, "JUNIO": 8, "JULIO": 9, "AGOSTO": 10,
    "SEPTIEMBRE": 11, "OCTUBRE": 12, "NOVIEMBRE": 13, "DICIEMBRE": 14
}

# Abreviatura usada en nombre de archivo → nombre completo del mes
MES_ABREV = {
    "ENE": "ENERO", "FEB": "FEBRERO", "MAR": "MARZO", "ABR": "ABRIL",
    "MAY": "MAYO", "JUN": "JUNIO", "JUL": "JULIO", "AGO": "AGOSTO",
    "SEP": "SEPTIEMBRE", "OCT": "OCTUBRE", "NOV": "NOVIEMBRE", "DIC": "DICIEMBRE"
}


# ---------------------------------------------------------------------------
# Utilidades
# ---------------------------------------------------------------------------
def celda_a_str(valor):
    """Convierte un valor de celda xlrd a string limpio."""
    if isinstance(valor, float) and valor == int(valor):
        return str(int(valor))
    return str(valor).strip()


def extraer_entidad(nombre_archivo):
    """
    Extrae el nombre de entidad del nombre del archivo.
    Ej: '5105_CHVS_FEB_2026-DL.xls'         → 'CHVS'
        '5105_UT ALIM YUMBO2026_FEB_2026.xls' → 'UT ALIM YUMBO2026'
    """
    patron = r'^5105_(.+?)_(ENE|FEB|MAR|ABR|MAY|JUN|JUL|AGO|SEP|OCT|NOV|DIC)_'
    match = re.search(patron, nombre_archivo, re.IGNORECASE)
    return match.group(1) if match else None


def extraer_mes_abrev(nombre_archivo):
    """
    Extrae la abreviatura del mes del nombre del archivo.
    Ej: '5105_CHVS_FEB_2026-DL.xls' → 'FEB'
    """
    patron = r'_(ENE|FEB|MAR|ABR|MAY|JUN|JUL|AGO|SEP|OCT|NOV|DIC)_'
    match = re.search(patron, nombre_archivo, re.IGNORECASE)
    return match.group(1).upper() if match else None


# ---------------------------------------------------------------------------
# Lógica principal de extracción por archivo
# ---------------------------------------------------------------------------
def procesar_archivo(filepath, reglas):
    """
    Abre un archivo XLS 5105 y calcula el valor según las reglas de la entidad.

    Reglas posibles en config:
        subtract_codes : list  → total5105 - suma(códigos indicados)
        only_code      : str   → toma directamente el valor del código indicado
    """
    nombre = os.path.basename(filepath)

    try:
        wb = xlrd.open_workbook(filepath)
        ws = wb.sheet_by_name('Hoja 1')
    except Exception as e:
        print(f"  [ERROR] No se pudo abrir {nombre}: {e}")
        return 0

    # --- 1. Detectar columna Débitos (busca el header "Débitos") -------------
    col_debitos = None
    for r in range(ws.nrows):
        for c in range(ws.ncols):
            if 'bito' in str(ws.cell_value(r, c)).lower():   # "Débitos" con o sin tilde
                col_debitos = c
                break
        if col_debitos is not None:
            break

    if col_debitos is None:
        print(f"  [ERROR] No se encontró columna Débitos en {nombre}")
        return 0

    # --- 2. Ubicar fila total 5105 bajo sección "99 GENERAL" ----------------
    en_seccion_99 = False
    fila_total = None
    filas_codigos = {}          # { "010303": fila_index, ... }

    for r in range(ws.nrows):
        a_val = celda_a_str(ws.cell_value(r, 0))

        # Detectar sección 99 GENERAL
        if '99' in a_val and 'GENERAL' in a_val.upper():
            en_seccion_99 = True
            continue

        if not en_seccion_99:
            continue

        # Primera fila con "5105" = total de la cuenta para U.N. 99
        if a_val == '5105' and fila_total is None:
            fila_total = r
            continue

        # A partir de ahí, recoger filas de detalle (códigos de 6 dígitos)
        if fila_total is not None:
            codigo_limpio = a_val.lstrip()
            match = re.match(r'^(\d{6})', codigo_limpio)
            if match:
                filas_codigos[match.group(1)] = r
            else:
                # Salimos al llegar a una nueva sección (header o línea no-código)
                if a_val:
                    break

    if fila_total is None:
        print(f"  [AVISO] No se encontró fila 5105/99-GENERAL en {nombre} → valor = 0")
        return 0

    # --- 3. Aplicar regla --------------------------------------------------
    if 'only_code' in reglas:
        # Tomar solo el valor del código indicado
        codigo = reglas['only_code']
        if codigo in filas_codigos:
            return ws.cell_value(filas_codigos[codigo], col_debitos) or 0
        print(f"  [AVISO] Código {codigo} no encontrado en {nombre} → valor = 0")
        return 0

    # Caso general: total - códigos a restar
    total = ws.cell_value(fila_total, col_debitos) or 0

    for codigo in reglas.get('subtract_codes', []):
        if codigo in filas_codigos:
            resta = ws.cell_value(filas_codigos[codigo], col_debitos) or 0
            total -= resta

    return total


# ---------------------------------------------------------------------------
# Función principal
# ---------------------------------------------------------------------------
def procesar_mes(carpeta_mes):
    """Procesa todos los archivos 5105 del mes y actualiza el xlsx principal."""

    # --- Detectar mes desde nombre de carpeta ---
    nombre_carpeta = os.path.basename(carpeta_mes.rstrip('/\\'))
    mes_nombre = nombre_carpeta.replace('-DL', '').upper()

    if mes_nombre not in MES_COLUMNA:
        print(f"ERROR: Mes '{mes_nombre}' no reconocido (carpeta: '{nombre_carpeta}')")
        print(f"       Carpetas válidas: ENERO-DL, FEBRERO-DL, MARZO-DL, ...")
        return

    columna_xlsx = MES_COLUMNA[mes_nombre]
    col_letra = chr(64 + columna_xlsx)
    print(f"{'='*60}")
    print(f"Mes: {mes_nombre}  |  Columna destino: {col_letra}")
    print(f"Carpeta: {carpeta_mes}")
    print(f"{'='*60}\n")

    # --- Cargar configuración ---
    with open(CONFIG_PATH, 'r', encoding='utf-8') as f:
        config = json.load(f)

    # --- Buscar archivos 5105_*.xls en la carpeta (sin recursividad) ---
    archivos = sorted([
        f for f in os.listdir(carpeta_mes)
        if re.match(r'^5105_', f, re.IGNORECASE)
        and f.lower().endswith('.xls')
        and os.path.isfile(os.path.join(carpeta_mes, f))
    ])

    if not archivos:
        print("No se encontraron archivos 5105_*.xls en la carpeta.")
        return

    print(f"Archivos encontrados: {len(archivos)}\n")

    total_general = 0
    resultados = []

    for nombre_archivo in archivos:
        entidad = extraer_entidad(nombre_archivo)

        if not entidad:
            print(f"  [OMITIDO] No se pudo extraer entidad de: {nombre_archivo}")
            continue

        if entidad not in config:
            print(f"  [OMITIDO] Sin reglas en config_5105.json → entidad='{entidad}'")
            continue

        reglas = config[entidad]
        filepath = os.path.join(carpeta_mes, nombre_archivo)
        valor = procesar_archivo(filepath, reglas)

        total_general += valor
        resultados.append((entidad, valor))
        print(f"  {entidad:<35}    ${valor:>18,.0f}")

    print(f"\n  {'TOTAL 5105':<35}   ${total_general:>18,.0f}")
    print()

    # --- Escribir en xlsx principal ---
    if not os.path.exists(ARCHIVO_PRINCIPAL):
        print(f"ERROR: No se encontró el archivo principal:\n  {ARCHIVO_PRINCIPAL}")
        return

    wb = openpyxl.load_workbook(ARCHIVO_PRINCIPAL)

    # Buscar hoja de Gastos Administrativos
    hoja = next(
        (wb[sh] for sh in wb.sheetnames if 'GASTO' in sh.upper() and 'ADMIN' in sh.upper()),
        None
    )
    if hoja is None:
        print("ERROR: No se encontró la hoja 'GASTOS ADMINISTRATIVOS' en el xlsx principal.")
        return

    # Buscar fila donde columna A = "5105"
    fila_destino = None
    for row in hoja.iter_rows():
        if str(row[0].value).strip() == '5105':
            fila_destino = row[0].row
            break

    if fila_destino is None:
        print("ERROR: No se encontró la fila con código '5105' en el xlsx principal.")
        return

    # Escribir valor
    celda = hoja.cell(row=fila_destino, column=columna_xlsx)
    valor_anterior = celda.value
    celda.value = total_general

    try:
        wb.save(ARCHIVO_PRINCIPAL)
        print(f"[OK] Celda {col_letra}{fila_destino} actualizada:")
        print(f"    Antes  : {valor_anterior}")
        print(f"    Ahora  : {total_general:,.0f}")
        print(f"    Archivo: {ARCHIVO_PRINCIPAL}")
    except PermissionError:
        print(f"[ERROR] No se pudo guardar el archivo.")
        print(f"        Cierre el archivo en Excel y vuelva a ejecutar el programa.")
        print(f"        Archivo: {ARCHIVO_PRINCIPAL}")


# ---------------------------------------------------------------------------
# Punto de entrada
# ---------------------------------------------------------------------------
if __name__ == '__main__':
    if len(sys.argv) > 1:
        arg = sys.argv[1]
        # Si es ruta absoluta la usa directamente, si no la construye desde BASE
        carpeta = arg if os.path.isabs(arg) else os.path.join(BASE, arg)
    else:
        carpeta = os.path.join(BASE, 'FEBRERO-DL')

    procesar_mes(carpeta)
