"""
procesar_todo.py
----------------
Procesa en un solo paso el archivo principal de Gastos Administrativos:
  1. Salarios (5105)   -> lee archivos 5105_*.xls  segun config_5105.json
  2. Gastos varios     -> lee archivos ER y 51355001 segun config_gastos.json

Uso:
    py procesar_todo.py                   # usa FEBRERO-DL por defecto
    py procesar_todo.py MARZO-DL          # nombre de carpeta relativo
    py procesar_todo.py C:\\ruta\\MARZO-DL # ruta absoluta
"""

import os, re, sys, json, warnings, unicodedata
from datetime import date
import xlrd, openpyxl
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter
import db as _db

warnings.filterwarnings('ignore')

# ---------------------------------------------------------------------------
BASE              = os.environ.get('CHVS_BASE', os.path.dirname(os.path.abspath(__file__)))
ARCHIVO_PRINCIPAL = os.path.join(BASE, 'INFORME REAL_2026 - FORMATO VERSION ORIGINAL.xlsx')

# Inicializar DB (por si el script se ejecuta directamente sin haber pasado por app.py)
_db.init_db()
_db.migrate_from_json()

# Columna por mes para hojas con 1 col por mes (ENE=3, FEB=4 ...)
MES_COLUMNA = {
    "ENERO":3,"FEBRERO":4,"MARZO":5,"ABRIL":6,"MAYO":7,"JUNIO":8,
    "JULIO":9,"AGOSTO":10,"SEPTIEMBRE":11,"OCTUBRE":12,"NOVIEMBRE":13,"DICIEMBRE":14
}
# Número ordinal del mes (para calcular columna en hojas con estructura diferente)
MES_NUMERO = {
    "ENERO":1,"FEBRERO":2,"MARZO":3,"ABRIL":4,"MAYO":5,"JUNIO":6,
    "JULIO":7,"AGOSTO":8,"SEPTIEMBRE":9,"OCTUBRE":10,"NOVIEMBRE":11,"DICIEMBRE":12
}

# Hojas con estructura de columnas diferente a la estándar (1 col/mes)
# Patrón 3 cols/mes: ENE=col3, FEB=col6, MAR=col9... → col = mes_num * 3
_tres_cols = lambda mes_num: mes_num * 3
COLUMNA_HOJA = {
    'CASINO': _tres_cols,
    'CALI'  : _tres_cols,
    'YUMBO' : _tres_cols,
    'BUGA'  : _tres_cols,
    'COMEDORES CALI'   : _tres_cols,
    'COMEDORES PALMIRA': _tres_cols,
    'COMEDORES VALLE'  : _tres_cols,
    'CTAS EN PPACION'  : _tres_cols,
    'PYG TOTAL'        : _tres_cols,
    'RECREARTE'        : _tres_cols,
}

MES_ABREV = {
    "ENE":"ENERO","FEB":"FEBRERO","MAR":"MARZO","ABR":"ABRIL",
    "MAY":"MAYO","JUN":"JUNIO","JUL":"JULIO","AGO":"AGOSTO",
    "SEP":"SEPTIEMBRE","OCT":"OCTUBRE","NOV":"NOVIEMBRE","DIC":"DICIEMBRE"
}

# ---------------------------------------------------------------------------
# Formulas cross-sheet internas para hojas de patron 3-col/mes (filas 148-151)
# c   = col GASTOS ADMINISTRATIVOS estandar  (get_column_letter(mes_num+2): ENE=C,FEB=D...)
# p   = col %DIST par  (get_column_letter(mes_num*2):   ENE=B,FEB=D,MAR=F...)
# im  = col %DIST impar (get_column_letter(mes_num*2+1): ENE=C,FEB=E,MAR=G...)
# k3  = col destino 3-col de las hojas referenciadas (get_column_letter(mes_num*3): ENE=C,FEB=F,MAR=I...)
# None = fila vacia en la plantilla, no se escribe formula
FORMULAS_CROSS_SHEET = {
    'CALI': {
        148: lambda c,p,im,k3: f"='GASTOS ADMINISTRATIVOS'!{c}96*'%DIST'!{p}36",
        149: lambda c,p,im,k3: f"='GASTOS ADMINISTRATIVOS'!{c}109*'%DIST'!{im}52+520117.4",
        150: None,
        151: lambda c,p,im,k3: f"='GASTOS ADMINISTRATIVOS'!{c}119*'%DIST'!{im}52",
    },
    'YUMBO': {
        148: lambda c,p,im,k3: f"='GASTOS ADMINISTRATIVOS'!{c}96*'%DIST'!{p}37",
        149: lambda c,p,im,k3: f"='GASTOS ADMINISTRATIVOS'!{c}109*'%DIST'!{im}53+29230.55+40203.2",
        150: None,
        151: lambda c,p,im,k3: f"='GASTOS ADMINISTRATIVOS'!{c}119*'%DIST'!{im}53",
    },
    'BUGA': {
        148: lambda c,p,im,k3: f"='GASTOS ADMINISTRATIVOS'!{c}96*'%DIST'!{p}39",
        149: lambda c,p,im,k3: f"='GASTOS ADMINISTRATIVOS'!{c}109*'%DIST'!{im}55+79552.5+3501.4",
        150: None,
        151: lambda c,p,im,k3: f"='GASTOS ADMINISTRATIVOS'!{c}119*'%DIST'!{im}55",
    },
    'COMEDORES CALI': {
        148: lambda c,p,im,k3: f"='GASTOS ADMINISTRATIVOS'!{c}96*'%DIST'!{p}41",
        149: lambda c,p,im,k3: f"='GASTOS ADMINISTRATIVOS'!{c}109*'%DIST'!{im}57",
        150: None,
        151: lambda c,p,im,k3: f"='GASTOS ADMINISTRATIVOS'!{c}119*'%DIST'!{im}57",
    },
    'COMEDORES PALMIRA': {
        148: lambda c,p,im,k3: f"='GASTOS ADMINISTRATIVOS'!{c}96*'%DIST'!{p}43",
        149: lambda c,p,im,k3: f"='GASTOS ADMINISTRATIVOS'!{c}109*'%DIST'!{im}59",
        150: None,
        151: lambda c,p,im,k3: f"='GASTOS ADMINISTRATIVOS'!{c}119*'%DIST'!{im}59",
    },
    'CTAS EN PPACION': {
        148: lambda c,p,im,k3: f"='GASTOS ADMINISTRATIVOS'!{c}96*'%DIST'!{p}45",
        149: None,
        150: None,
        151: None,
    },
    'PYG TOTAL': {
        # Suma de todas las unidades; referencia col 3-col (k3) de cada hoja origen
        148: lambda c,p,im,k3: (
            f"=CALI!{k3}148+HUNGRIA!{k3}148+YUMBO!{k3}148+'ADULTO MAYOR'!{k3}148"
            f"+'COMEDORES VALLE'!{k3}148+BUGA!{k3}148+'COMEDORES CALI'!{k3}148"
            f"+'CTAS EN PPACION'!{k3}148+'COMEDORES PALMIRA'!{k3}148+IMDERTY!{k3}148"
        ),
        149: lambda c,p,im,k3: (
            f"=CALI!{k3}149+HUNGRIA!{k3}149+YUMBO!{k3}149+'ADULTO MAYOR'!{k3}149"
            f"+'COMEDORES VALLE'!{k3}149+BUGA!{k3}149+'COMEDORES CALI'!{k3}149"
            f"+'CTAS EN PPACION'!{k3}149+'COMEDORES PALMIRA'!{k3}149+IMDERTY!{k3}149"
        ),
        150: lambda c,p,im,k3: (
            f"=CALI!{k3}150+HUNGRIA!{k3}150+YUMBO!{k3}150+'ADULTO MAYOR'!{k3}150"
            f"+'COMEDORES VALLE'!{k3}150+BUGA!{k3}150+'COMEDORES CALI'!{k3}150"
            f"+'CTAS EN PPACION'!{k3}150+'COMEDORES PALMIRA'!{k3}150+IMDERTY!{k3}150"
        ),
        151: lambda c,p,im,k3: (
            f"=CALI!{k3}151+HUNGRIA!{k3}151+YUMBO!{k3}151+'ADULTO MAYOR'!{k3}151"
            f"+'COMEDORES VALLE'!{k3}151+BUGA!{k3}151+'COMEDORES CALI'!{k3}151"
            f"+'CTAS EN PPACION'!{k3}151+'COMEDORES PALMIRA'!{k3}151+IMDERTY!{k3}151"
        ),
    },
}


# ---------------------------------------------------------------------------
# Utilidades comunes
# ---------------------------------------------------------------------------
def celda_str(raw):
    """Convierte valor de celda xlrd a string limpio (maneja float como 51956001.0)."""
    if isinstance(raw, float) and raw == int(raw):
        return str(int(raw))
    return str(raw).strip()


def norm_text(s):
    """Normaliza texto para detección robusta de encabezados."""
    t = str(s or '').lower().replace('ã©', 'e')
    t = unicodedata.normalize('NFKD', t)
    return ''.join(ch for ch in t if not unicodedata.combining(ch))


def encontrar_archivo(carpeta, prefijo, entidad, mes_abrev, anio):
    """Busca archivo con patron: {prefijo}_{entidad}_{mes}_{anio}*.xls
    1. Entidad exacta + año exacto
    2. Entidad exacta + cualquier año de 4 dígitos
    3. Entidad como prefijo (permite dígitos pegados, ej. 'CONS ALIM CALI' → 'CONS ALIM CALI2026')
    """
    p = rf'^{re.escape(prefijo)}_{re.escape(entidad)}'
    patron_exact   = re.compile(p + rf'_{mes_abrev}_{anio}.*\.xls$',  re.IGNORECASE)
    patron_flex    = re.compile(p + rf'_\d{{4}}.*\.xls$',              re.IGNORECASE)
    # ponytail: \d* permite año pegado a entidad (CONS ALIM CALI2026_MAY_*)
    patron_prefix  = re.compile(p + rf'\d*_{mes_abrev}_\d{{4}}.*\.xls$', re.IGNORECASE)
    archivos = sorted(os.listdir(carpeta))
    for f in archivos:
        if patron_exact.match(f) and os.path.isfile(os.path.join(carpeta, f)):
            return os.path.join(carpeta, f)
    for f in archivos:
        if patron_flex.match(f) and os.path.isfile(os.path.join(carpeta, f)):
            return os.path.join(carpeta, f)
    matches = [f for f in archivos
               if patron_prefix.match(f) and os.path.isfile(os.path.join(carpeta, f))]
    if matches:
        return os.path.join(carpeta, sorted(matches)[-1])  # más reciente alfabéticamente
    return None


def detectar_col_debitos(ws, max_filas=25):
    """Detecta la columna Debitos buscando el encabezado en las primeras filas."""
    for r in range(min(max_filas, ws.nrows)):
        for c in range(ws.ncols):
            txt = norm_text(ws.cell_value(r, c))
            if 'netos' in txt or 'debito' in txt:
                return c
    return 9  # fallback columna J


def detectar_col_valor(ws, max_filas=25):
    """
    Retorna (columna, modo):
      - modo='debito' cuando encuentra encabezado Debitos
      - modo='netos'  cuando encuentra Netos y no hay Debitos
    """
    col_netos = None
    for r in range(min(max_filas, ws.nrows)):
        for c in range(ws.ncols):
            txt = norm_text(ws.cell_value(r, c))
            if 'debito' in txt:
                return c, 'debito'
            if 'netos' in txt and col_netos is None:
                col_netos = c
    if col_netos is not None:
        return col_netos, 'netos'
    return 9, 'debito'


def detectar_col_creditos(ws, max_filas=25):
    """Detecta la columna Creditos buscando el encabezado en las primeras filas."""
    for r in range(min(max_filas, ws.nrows)):
        for c in range(ws.ncols):
            txt = norm_text(ws.cell_value(r, c))
            if 'credito' in txt:
                return c
    return None


# ---------------------------------------------------------------------------
# Lectura ESTADO DE RESULTADOS
# ---------------------------------------------------------------------------
def leer_er(filepath, codigos_buscar, sin_filtro=False, seccion=None):
    """
    Suma los debitos de los codigos indicados dentro de la seccion U.N. 99 GENERAL.
    Si sin_filtro=True, busca en todo el archivo.
    """
    wb = xlrd.open_workbook(filepath)
    ws = wb.sheet_by_name('Hoja 1')
    col_val, modo_valor = detectar_col_valor(ws)
    col_credito = detectar_col_creditos(ws)

    inicio_99 = None
    fin_99    = ws.nrows

    # Si viene seccion explicita del chip/config, usar ese bloque.
    if seccion:
        en_seccion = False
        inicio_99 = ws.nrows
        for r in range(ws.nrows):
            a_raw = str(ws.cell_value(r, 0))
            a_str = a_raw.strip()
            if a_str == seccion:
                inicio_99 = r + 1
                en_seccion = True
                continue
            if en_seccion:
                # Nueva cabecera de seccion: texto sin sangria, no codigo numerico puro
                if (a_raw and not a_raw[0].isspace() and a_str
                        and not re.match(r'^\d+(\.0)?$', a_str)):
                    fin_99 = r
                    break
        if inicio_99 == ws.nrows:
            inicio_99, fin_99 = 0, ws.nrows
    else:
        # Modo legado: U.N. 99 GENERAL
        for r in range(ws.nrows):
            a = str(ws.cell_value(r, 0)).strip()
            if inicio_99 is None:
                if '99' in a and 'GENERAL' in a.upper():
                    inicio_99 = r + 1
            else:
                if a and not re.match(r'^[\d\s]', a) and a not in ('Cuentas',):
                    fin_99 = r
                    break

    if sin_filtro:
        inicio_99, fin_99 = 0, ws.nrows
    elif inicio_99 is None:
        inicio_99 = 0

    col_letra = chr(65 + col_val)
    total     = 0.0
    celdas    = []
    for r in range(inicio_99, fin_99):
        if celda_str(ws.cell_value(r, 0)) in codigos_buscar:
            deb = float(ws.cell_value(r, col_val) or 0)
            usa_resta_jk = (modo_valor == 'debito' and col_credito is not None)
            cre = float(ws.cell_value(r, col_credito) or 0) if usa_resta_jk else 0.0
            v = (deb - cre) if usa_resta_jk else deb
            total += v
            celdas.append(f"{col_letra}{r + 1}")
    return total, celdas


# ---------------------------------------------------------------------------
# Lectura archivos auxiliares (51355001)
# ---------------------------------------------------------------------------
def leer_aux(filepath, codigos_buscar, seccion=None):
    """
    Suma debitos de filas cuya col A coincida exactamente con el codigo
    o comience con el codigo seguido de espacio (filas NIT + nombre).
    Si seccion no es None y el archivo es un 7205, busca solo dentro
    de esa sección (entre su cabecera y la siguiente cabecera de sección).
    """
    wb = xlrd.open_workbook(filepath)
    ws = wb.sheet_by_name('Hoja 1')
    col_val, modo_valor = detectar_col_valor(ws)
    col_credito = detectar_col_creditos(ws)

    nombre       = os.path.basename(filepath).upper()
    es_seccionado = nombre.startswith('7205_') or nombre.startswith('7105_')

    # ponytail: códigos 5105 pueden venir en formato antiguo "010401   PLANEACION"
    # → normalizar a solo la parte numérica para que startswith funcione
    if nombre.startswith('5105_'):
        codigos_buscar = {
            (re.match(r'^(\d+)\s+', c).group(1) if re.match(r'^(\d+)\s+', c) else c)
            for c in codigos_buscar
        }

    # Rango de filas a buscar (todo el archivo por defecto)
    fila_min, fila_max = 0, ws.nrows

    if seccion and es_seccionado:
        # Localizar el inicio de la sección y el inicio de la siguiente
        en_seccion = False
        fila_min   = ws.nrows  # se actualiza al encontrar la sección
        for r in range(ws.nrows):
            a_raw = str(ws.cell_value(r, 0))
            a_str = a_raw.strip()
            if a_str == seccion:
                fila_min   = r + 1
                en_seccion = True
                continue
            if en_seccion:
                # Nueva cabecera de sección: sin espacio inicial, no es código puro
                if (a_raw and not a_raw[0].isspace() and a_str
                        and not re.match(r'^\d{3,8}$', a_str)):
                    fila_max = r
                    break

    col_letra = chr(65 + col_val)
    total     = 0.0
    celdas    = []
    for r in range(fila_min, fila_max):
        a_val = celda_str(ws.cell_value(r, 0))
        # Códigos con descripción (contienen espacio): solo exact match.
        # Códigos puramente numéricos: también acepta startswith para capturar
        # filas del tipo "029903 NIT12345" cuando el código buscado es "029903".
        matched = a_val in codigos_buscar or any(
            (' ' not in code) and (
                a_val.startswith(code + ' ') or a_val.startswith(code + '\t')
            )
            for code in codigos_buscar
        )
        if matched:
            deb = float(ws.cell_value(r, col_val) or 0)
            usa_resta_jk = (modo_valor == 'debito' and col_credito is not None)
            cre = float(ws.cell_value(r, col_credito) or 0) if usa_resta_jk else 0.0
            v = (deb - cre) if usa_resta_jk else deb
            total += v
            celdas.append(f"{col_letra}{r + 1}")
    return total, celdas


# ---------------------------------------------------------------------------
# Lectura archivos 5105
# ---------------------------------------------------------------------------
def leer_5105(filepath, reglas):
    """
    Abre un archivo 5105 y calcula el valor segun las reglas de la entidad:
      subtract_codes : total5105 - suma de los codigos indicados
      only_code      : solo el valor del codigo indicado
    """
    wb = xlrd.open_workbook(filepath)
    ws = wb.sheet_by_name('Hoja 1')
    col_debitos, modo_valor = detectar_col_valor(ws, max_filas=ws.nrows)
    col_creditos = detectar_col_creditos(ws, max_filas=ws.nrows)
    usa_resta_jk = (modo_valor == 'debito' and col_creditos is not None)

    en_seccion_99 = False
    fila_total    = None
    filas_codigos = {}

    for r in range(ws.nrows):
        a_val = celda_str(ws.cell_value(r, 0))

        if '99' in a_val and 'GENERAL' in a_val.upper():
            en_seccion_99 = True
            continue
        if not en_seccion_99:
            continue

        if a_val == '5105' and fila_total is None:
            fila_total = r
            continue

        if fila_total is not None:
            match = re.match(r'^(\d{6})', a_val.lstrip())
            if match:
                filas_codigos[match.group(1)] = r
            elif a_val:
                break

    col_letra = chr(65 + col_debitos)

    if fila_total is None:
        return 0, []

    if 'only_code' in reglas:
        codigo = reglas['only_code']
        if codigo in filas_codigos:
            r = filas_codigos[codigo]
            v = (ws.cell_value(r, col_debitos) or 0) - (
                (ws.cell_value(r, col_creditos) or 0) if usa_resta_jk else 0
            )
            detalle = [f"  Codigo {codigo}: ${v:,.0f}  (celda {col_letra}{r+1})"]
            return v, detalle
        return 0, []

    total = (ws.cell_value(fila_total, col_debitos) or 0) - (
        (ws.cell_value(fila_total, col_creditos) or 0) if usa_resta_jk else 0
    )
    detalle = [f"  Total 5105: ${total:,.0f}  (celda {col_letra}{fila_total+1})"]
    for codigo in reglas.get('subtract_codes', []):
        if codigo in filas_codigos:
            r   = filas_codigos[codigo]
            v   = (ws.cell_value(r, col_debitos) or 0) - (
                (ws.cell_value(r, col_creditos) or 0) if usa_resta_jk else 0
            )
            total -= v
            detalle.append(f"  - {codigo}: ${v:,.0f}  (celda {col_letra}{r+1})")
    detalle.append(f"  Neto: ${total:,.0f}")
    return total, detalle


# ---------------------------------------------------------------------------
# PASO 1: Salarios 5105
# ---------------------------------------------------------------------------
def procesar_5105(carpeta_mes, hoja_dest, columna_xlsx, mes_nombre):
    print(f"\n{'='*65}")
    print(f"PASO 1 - SALARIOS (5105)")
    print(f"{'='*65}")

    config = _db.load_config_5105()

    archivos = sorted([
        f for f in os.listdir(carpeta_mes)
        if re.match(r'^5105_', f, re.IGNORECASE)
        and f.lower().endswith('.xls')
        and os.path.isfile(os.path.join(carpeta_mes, f))
    ])

    if not archivos:
        print("  No se encontraron archivos 5105_*.xls")
        return

    total_general    = 0.0
    detalle_entidades = []
    fuentes_log       = []
    for nombre in archivos:
        patron = r'^5105_(.+?)_(ENE|FEB|MAR|ABR|MAY|JUN|JUL|AGO|SEP|OCT|NOV|DIC)_'
        m = re.search(patron, nombre, re.IGNORECASE)
        entidad = m.group(1) if m else None

        if not entidad or entidad not in config:
            print(f"  [OMITIDO] {nombre}")
            continue

        valor, detalle = leer_5105(os.path.join(carpeta_mes, nombre), config[entidad])
        total_general += valor
        detalle_entidades.append(f"Archivo: {nombre}\n" + "\n".join(detalle))
        fuentes_log.append({'file_key': f'5105_{entidad}', 'codigo': '5105',
                            'seccion': '99 GENERAL', 'op': '+', 'valor': valor})
        print(f"  {entidad:<40} ${valor:>16,.0f}")

    print(f"  {'TOTAL 5105':<40} ${total_general:>16,.0f}")

    fila_escrita     = None
    valor_a_escribir = total_general  # se actualiza con valor_fijo al encontrar la fila
    # Escribir en destino con comentario de trazabilidad
    for row in hoja_dest.iter_rows():
        if str(row[0].value or '').strip() == '5105':
            fila_5105, col_5105 = row[0].row, columna_xlsx
            for rango in hoja_dest.merged_cells.ranges:
                if (rango.min_row <= fila_5105 <= rango.max_row and
                        rango.min_col <= col_5105 <= rango.max_col):
                    fila_5105 = rango.min_row
                    col_5105  = rango.min_col
                    break
            celda = hoja_dest.cell(row=fila_5105, column=col_5105)
            valor_a_escribir = total_general
            cfg_items = _db.load_config(mes_nombre).get('items', [])
            cfg_5105 = next((c for c in cfg_items
                             if c.get('hoja', '').strip() == 'GASTOS ADMINISTRATIVOS'
                             and c.get('codigo_a') == '5105'), None)
            if cfg_5105:
                for op_item in (cfg_5105.get('valor_fijo') or []):
                    op, val = op_item.get('op', '+'), float(op_item.get('valor', 0))
                    if   op == '+': valor_a_escribir += val
                    elif op == '-': valor_a_escribir -= val
                    elif op == '*': valor_a_escribir *= val
                    elif op == '/' and val: valor_a_escribir /= val
            celda.value = valor_a_escribir
            texto = "\n".join(detalle_entidades)
            texto += f"\n{'-'*30}\nTotal: ${valor_a_escribir:,.0f}\nGenerado: {date.today()}"
            comentario = Comment(texto, "Script")
            comentario.width  = 350
            comentario.height = 80 + 20 * len(detalle_entidades)
            celda.comment = comentario
            fila_escrita = fila_5105
            print(f"  -> Escrito en fila {row[0].row}, columna {chr(64+columna_xlsx)}")
            break

    return {
        'hoja'       : 'GASTOS ADMINISTRATIVOS',
        'fila'       : fila_escrita or 0,
        'codigo_a'   : '5105',
        'valor_total': valor_a_escribir,
        'fuentes'    : fuentes_log,
    }


# ---------------------------------------------------------------------------
# Helper: construir mapa key → ruta de archivo
# ---------------------------------------------------------------------------
_RE_MES_STRIP = re.compile(
    r'_(ENE|FEB|MAR|ABR|MAY|JUN|JUL|AGO|SEP|OCT|NOV|DIC'
    r'|ENERO|FEBRERO|MARZO|ABRIL|MAYO|JUNIO|JULIO|AGOSTO'
    r'|SEPTIEMBRE|OCTUBRE|NOVIEMBRE|DICIEMBRE)_\d{4}$',
    re.IGNORECASE
)

def _construir_archivos_cache(carpeta_mes, file_sources, mes_abrev, anio):
    """Construye {key: ruta} con file_sources + auto-discovery + alias sin año."""
    cache = {}
    for key, info in file_sources.items():
        cache[key] = encontrar_archivo(carpeta_mes, info['prefijo'], info['entidad'], mes_abrev, anio)

    for _fname in sorted(os.listdir(carpeta_mes)):
        if not _fname.lower().endswith('.xls'):
            continue
        _fpath = os.path.join(carpeta_mes, _fname)
        if not os.path.isfile(_fpath):
            continue
        _base = _RE_MES_STRIP.sub('', os.path.splitext(_fname)[0])
        _auto_key = re.sub(r'\s+', ' ', _base).strip().replace(' ', '_')
        if not cache.get(_auto_key):
            cache[_auto_key] = _fpath

    _keys_con_ruta = [k for k, v in cache.items() if v]
    for _src_key in list(cache.keys()):
        if cache.get(_src_key):
            continue
        _matches = sorted([k for k in _keys_con_ruta if k.startswith(_src_key)])
        if _matches:
            cache[_src_key] = cache[_matches[-1]]

    return cache


# ---------------------------------------------------------------------------
# Preview: calcular valores sin escribir en el Excel
# ---------------------------------------------------------------------------
def calcular_preview(carpeta_mes):
    """
    Calcula todos los valores que se escribirían sin modificar el Excel.
    Retorna lista de {hoja, fila, codigo_a, valor_total, advertencias, fuentes}.
    """
    nombre_carpeta = os.path.basename(carpeta_mes.rstrip('/\\'))
    mes_nombre     = nombre_carpeta.replace('-DL', '').upper()
    if mes_nombre not in MES_COLUMNA:
        raise ValueError(f"Mes '{mes_nombre}' no reconocido.")

    mes_abrev = next((k for k, v in MES_ABREV.items() if v == mes_nombre), None)
    anio = None
    for f in os.listdir(carpeta_mes):
        m = re.search(r'_(\d{4})[-\.]', f)
        if m:
            anio = m.group(1)
            break
    if not anio:
        raise ValueError("No se pudo detectar el año desde los nombres de archivo.")

    config       = _db.load_config(mes_nombre)
    file_sources = config['file_sources']
    items        = config['items']
    cache        = _construir_archivos_cache(carpeta_mes, file_sources, mes_abrev, anio)

    resultado = []
    for item in items:
        codigo_a  = item.get('codigo_a', '')
        hoja      = item.get('hoja', 'GASTOS ADMINISTRATIVOS').strip()
        fila_dest = item.get('fila_dest')

        valor_total  = 0.0
        advertencias = []
        fuentes_log  = []

        for src in item.get('sources', []):
            key   = src['key']
            codes = src.get('codes', [])
            ruta  = cache.get(key)
            if not ruta:
                advertencias.append(f"NO ENCONTRADO: {key}")
                continue
            try:
                prefijo = file_sources.get(key, {}).get('prefijo') or (
                    'ESTADO DE RESULTADOS'
                    if os.path.basename(ruta).upper().startswith('ESTADO') else 'AUX'
                )
                if prefijo == 'ESTADO DE RESULTADOS':
                    v, _ = leer_er(ruta, set(codes),
                                   sin_filtro=src.get('sin_filtro', False),
                                   seccion=src.get('seccion'))
                else:
                    v, _ = leer_aux(ruta, set(codes), seccion=src.get('seccion'))
                op = src.get('op', '+')
                if   op == '+': valor_total += v
                elif op == '-': valor_total -= v
                elif op == '*': valor_total *= v
                elif op == '/': valor_total = valor_total / v if v != 0 else valor_total
                fuentes_log.append({'file_key': key, 'codigo': ', '.join(codes),
                                    'seccion': src.get('seccion'), 'op': op, 'valor': v})
            except Exception as e:
                advertencias.append(f"ERROR {key}: {e}")

        valor_fijo = item.get('valor_fijo')
        if valor_fijo:
            ops = valor_fijo if isinstance(valor_fijo, list) else [valor_fijo]
            for op_item in ops:
                op  = op_item.get('op', '+')
                val = float(op_item.get('valor', 0))
                if   op == '+': valor_total += val
                elif op == '-': valor_total -= val
                elif op == '*': valor_total *= val
                elif op == '/': valor_total = valor_total / val if val != 0 else valor_total

        if fila_dest or valor_total != 0 or advertencias:
            resultado.append({
                'hoja'       : hoja,
                'fila'       : fila_dest,
                'codigo_a'   : codigo_a,
                'valor_total': valor_total,
                'advertencias': advertencias,
                'fuentes'    : fuentes_log,
            })


    return resultado


# ---------------------------------------------------------------------------
# PASO 2: Gastos varios
# ---------------------------------------------------------------------------
def procesar_gastos(carpeta_mes, wb_dest, columna_xlsx, mes_abrev, anio, mes_nombre=None):
    print(f"\n{'='*65}")
    print(f"PASO 2 - GASTOS (todas las hojas)")
    print(f"{'='*65}")

    config = _db.load_config(mes_nombre)

    file_sources = config['file_sources']
    items        = config['items']
    valores_log  = []

    archivos_cache = _construir_archivos_cache(carpeta_mes, file_sources, mes_abrev, anio)

    for key, ruta in archivos_cache.items():
        if key in file_sources:
            estado = "OK" if ruta else "NO ENCONTRADO"
            print(f"  [{estado}] {key:<20} -> {os.path.basename(ruta) if ruta else '-'}")

    print()

    def normalizar(texto):
        return re.sub(r'\s+', ' ', str(texto).strip())

    # Índice de hojas: nombre normalizado (sin espacios) → nombre real en el libro
    hojas_idx = {sh.strip(): sh for sh in wb_dest.sheetnames}

    # Agrupar items por hoja destino (siempre con nombre normalizado como clave)
    from collections import defaultdict
    items_por_hoja = defaultdict(list)
    for item in items:
        hoja_nombre = item.get('hoja', '').strip()
        if not hoja_nombre:
            # Compatibilidad: items sin hoja → GASTOS ADMINISTRATIVOS
            hoja_nombre = next(
                (sh.strip() for sh in wb_dest.sheetnames if 'GASTO' in sh.upper() and 'ADMIN' in sh.upper()),
                wb_dest.sheetnames[0].strip()
            )
        items_por_hoja[hoja_nombre].append(item)

    # Limpiar columna del mes en TODAS las hojas antes de escribir.
    # Así, si una hoja queda sin items en el config (cero chips), su columna
    # también se limpia y no quedan residuos de ejecuciones anteriores.
    # Solo borra valores numéricos y fórmulas cross-sheet (contienen '!').
    # Las fórmulas internas de totales (=SUM(...)) se preservan.
    mes_num_global = MES_NUMERO.get(mes_nombre or '', 0)
    for hoja_real_all in wb_dest.sheetnames:
        ws_clear = wb_dest[hoja_real_all]
        hoja_norm_all = hoja_real_all.strip()
        if hoja_norm_all in COLUMNA_HOJA and mes_num_global:
            col_clear = COLUMNA_HOJA[hoja_norm_all](mes_num_global)
        else:
            col_clear = columna_xlsx
        for fila_row in ws_clear.iter_rows(min_col=col_clear, max_col=col_clear):
            cell = fila_row[0]
            if cell.value is None:
                continue
            es_numero  = isinstance(cell.value, (int, float))
            es_cs_form = isinstance(cell.value, str) and cell.value.startswith('=') and '!' in cell.value
            if es_numero or es_cs_form:
                cell.value   = None
                cell.comment = None

    # Procesar cada hoja
    for hoja_nombre, hoja_items in items_por_hoja.items():
        hoja_real = hojas_idx.get(hoja_nombre)
        if hoja_real is None:
            print(f"  [ADVERTENCIA] Hoja no encontrada: {hoja_nombre}")
            continue

        hoja_dest = wb_dest[hoja_real]

        # Calcular columna destino según estructura de la hoja
        mes_num = MES_NUMERO.get(mes_nombre or '', 0)
        if hoja_nombre.strip() in COLUMNA_HOJA and mes_num:
            col_hoja = COLUMNA_HOJA[hoja_nombre.strip()](mes_num)
        else:
            col_hoja = columna_xlsx

        print(f"\n  --- {hoja_real.strip()} ({len(hoja_items)} items) col={col_hoja} ---")

        # Construir índices de fila
        indice_destino   = {}
        indice_destino_b = {}
        for row in hoja_dest.iter_rows():
            cod  = normalizar(row[0].value) if row[0].value else ''
            desc = normalizar(row[1].value) if len(row) > 1 and row[1].value else ''
            if cod:
                indice_destino[cod] = row[0].row
            if desc:
                indice_destino_b[desc] = row[0].row

        actualizados   = 0
        no_encontrados = []

        print(f"  {'Codigo destino':<32} {'Valor calculado':>18}   Fila")
        print(f"  {'-'*65}")

        for item in hoja_items:
            codigo_a          = item['codigo_a']
            valor_total       = 0.0
            advertencias      = []
            lineas_comentario = []
            fuentes_item_log  = []

            for src in item.get('sources', []):
                key   = src['key']
                codes = src.get('codes', [])
                ruta  = archivos_cache.get(key)
                if not ruta:
                    advertencias.append(f"[NO ENCONTRADO: {key}]")
                    continue
                try:
                    if key in file_sources:
                        prefijo = file_sources[key]['prefijo']
                    else:
                        prefijo = ('ESTADO DE RESULTADOS'
                                   if os.path.basename(ruta).upper().startswith('ESTADO')
                                   else 'AUX')
                    if prefijo == 'ESTADO DE RESULTADOS':
                        v, celdas = leer_er(
                            ruta,
                            set(codes),
                            sin_filtro=src.get('sin_filtro', False),
                            seccion=src.get('seccion')
                        )
                    else:
                        v, celdas = leer_aux(ruta, set(codes), seccion=src.get('seccion'))
                    op_src = src.get('op', '+')
                    if   op_src == '+': valor_total += v
                    elif op_src == '-': valor_total -= v
                    elif op_src == '*': valor_total *= v
                    elif op_src == '/': valor_total = valor_total / v if v != 0 else valor_total
                    fuentes_item_log.append({
                        'file_key': key,
                        'codigo'  : ', '.join(codes),
                        'seccion' : src.get('seccion'),
                        'op'      : op_src,
                        'valor'   : v,
                    })
                    lineas_comentario.append(
                        f"Archivo: {os.path.basename(ruta)}\n"
                        f"  Op:      {op_src}\n"
                        f"  Codigos: {', '.join(codes)}\n"
                        f"  Celdas:  {', '.join(celdas) if celdas else 'no encontrado'}\n"
                        f"  Valor:   ${v:,.0f}"
                    )
                except Exception as e:
                    advertencias.append(f"[ERROR {key}: {e}]")

            # Operaciones fijas hardcodeadas
            valor_fijo = item.get('valor_fijo')
            if valor_fijo:
                ops = valor_fijo if isinstance(valor_fijo, list) else [valor_fijo]
                for op_item in ops:
                    op  = op_item.get('op', '+')
                    val = float(op_item.get('valor', 0))
                    if val == 0:
                        continue
                    if   op == '+': valor_total = valor_total + val
                    elif op == '-': valor_total = valor_total - val
                    elif op == '*': valor_total = valor_total * val
                    elif op == '/': valor_total = valor_total / val if val != 0 else valor_total
                    lineas_comentario.append(f"Op fija: {op} ${val:,.0f}  -> subtotal ${valor_total:,.0f}")

            # Buscar fila destino
            # 1) fila_dest: lookup directo por número de fila (evita colisiones de código duplicado)
            fila_dest  = item.get('fila_dest')
            buscar_por = item.get('buscar_por', 'A')
            if fila_dest:
                fila = fila_dest
            elif buscar_por == 'B':
                desc_b = normalizar(codigo_a[2:] if codigo_a.startswith('B:') else codigo_a)
                fila = indice_destino_b.get(desc_b)
            else:
                fila = indice_destino.get(normalizar(codigo_a))

            if fila is None:
                no_encontrados.append(codigo_a)
                print(f"  {codigo_a:<32} [SIN FILA DESTINO]")
                continue

            # Resolver celda real (puede ser merged — solo la top-left es escribible)
            fila_real, col_real = fila, col_hoja
            for rango in hoja_dest.merged_cells.ranges:
                if (rango.min_row <= fila <= rango.max_row and
                        rango.min_col <= col_hoja <= rango.max_col):
                    fila_real = rango.min_row
                    col_real  = rango.min_col
                    break

            # Escribir valor y comentario
            celda = hoja_dest.cell(row=fila_real, column=col_real)
            celda.value = valor_total
            texto_comentario  = "\n".join(lineas_comentario)
            texto_comentario += f"\n{'-'*30}\nTotal: ${valor_total:,.0f}\nGenerado: {date.today()}"
            if advertencias:
                texto_comentario += "\n" + "\n".join(advertencias)
            comentario = Comment(texto_comentario, "Script")
            comentario.width  = 300
            comentario.height = 100 + 20 * len(lineas_comentario)
            celda.comment = comentario

            # Guardar en log de ejecución
            valores_log.append({
                'hoja'       : hoja_nombre,
                'fila'       : fila_real,
                'codigo_a'   : codigo_a,
                'valor_total': valor_total,
                'fuentes'    : fuentes_item_log,
            })

            actualizados += 1
            adv = '  ' + ' '.join(advertencias) if advertencias else ''
            print(f"  {codigo_a:<32} {valor_total:>18,.2f}   fila {fila}{adv}")

        print(f"\n  Items actualizados : {actualizados}")
        if no_encontrados:
            print(f"  Sin fila destino   : {no_encontrados}")

    return valores_log


# ---------------------------------------------------------------------------
# Funcion principal
# ---------------------------------------------------------------------------
def procesar_mes(carpeta_mes):
    nombre_carpeta = os.path.basename(carpeta_mes.rstrip('/\\'))
    mes_nombre     = nombre_carpeta.replace('-DL', '').upper()

    if mes_nombre not in MES_COLUMNA:
        print(f"ERROR: Mes '{mes_nombre}' no reconocido.")
        return

    mes_abrev = next((k for k, v in MES_ABREV.items() if v == mes_nombre), None)
    if not mes_abrev:
        print(f"ERROR: No se pudo determinar abreviatura para '{mes_nombre}'")
        return

    anio = None
    for f in os.listdir(carpeta_mes):
        m = re.search(r'_(\d{4})[-\.]', f)
        if m:
            anio = m.group(1)
            break
    if not anio:
        print("ERROR: No se pudo detectar el anio desde los nombres de archivo.")
        return

    columna_xlsx = MES_COLUMNA[mes_nombre]
    col_letra    = chr(64 + columna_xlsx)

    print(f"\n{'#'*65}")
    print(f"  Mes: {mes_nombre} | Anio: {anio} | Columna destino: {col_letra}")
    print(f"  Carpeta: {carpeta_mes}")
    print(f"{'#'*65}")

    if not os.path.exists(ARCHIVO_PRINCIPAL):
        print(f"ERROR: Archivo principal no encontrado:\n  {ARCHIVO_PRINCIPAL}")
        return

    wb_dest   = openpyxl.load_workbook(ARCHIVO_PRINCIPAL)

    hoja_admin = next(
        (wb_dest[sh] for sh in wb_dest.sheetnames
         if 'GASTO' in sh.upper() and 'ADMIN' in sh.upper()),
        None
    )

    # Paso 1: Gastos (todas las hojas según config)
    log_gastos = procesar_gastos(carpeta_mes, wb_dest, columna_xlsx, mes_abrev, anio, mes_nombre)


    # Paso 3: Fórmulas internas cross-sheet en GASTOS ADMINISTRATIVOS
    # F89  = CASINO!{col_casino}$144  (col_casino = mes_num * 3)
    # F114 = RECREARTE!{col_recrearte}$141  (col_recrearte = mes_num * 3)
    mes_num = MES_NUMERO.get(mes_nombre, 0)
    if mes_num and hoja_admin:
        col_3 = chr(64 + mes_num * 3)   # columna en hojas de patrón 3-cols/mes
        hoja_admin.cell(row=89,  column=columna_xlsx).value = f'=CASINO!${col_3}$144'
        hoja_admin.cell(row=114, column=columna_xlsx).value = f'=RECREARTE!${col_3}$141'
        print(f"[OK] Formulas internas escritas: F89=CASINO!{col_3}144, F114=RECREARTE!{col_3}141")

    # Paso 4: Formulas cross-sheet en hojas 3-col/mes (filas 148-151)
    # col_adm  = col GASTOS ADMINISTRATIVOS estandar (ENE=C, FEB=D, MAR=E...)
    # col_par  = col %DIST par   (ENE=B, FEB=D, MAR=F...)
    # col_imp  = col %DIST impar (ENE=C, FEB=E, MAR=G...)
    # col3     = col destino 3-col de hojas referenciadas (ENE=C, FEB=F, MAR=I...)
    # col_dest = numero de columna destino en la hoja (ENE=3, FEB=6, MAR=9...)
    if mes_num:
        col_adm  = get_column_letter(mes_num + 2)      # ENE=C, FEB=D, MAR=E...
        col_par  = get_column_letter(mes_num * 2)      # ENE=B, FEB=D, MAR=F...
        col_imp  = get_column_letter(mes_num * 2 + 1)  # ENE=C, FEB=E, MAR=G...
        col3     = get_column_letter(mes_num * 3)      # ENE=C, FEB=F, MAR=I...
        col_dest = mes_num * 3                         # ENE=3, FEB=6, MAR=9...

        for hoja_nombre, filas_config in FORMULAS_CROSS_SHEET.items():
            if hoja_nombre not in wb_dest.sheetnames:
                continue
            ws = wb_dest[hoja_nombre]
            for fila, formula_fn in filas_config.items():
                if formula_fn is None:
                    continue
                formula = formula_fn(col_adm, col_par, col_imp, col3)
                ws.cell(row=fila, column=col_dest).value = formula
                print(f"[OK] {hoja_nombre} fila {fila} col {col_dest} = {formula}")

    # Forzar recálculo automático al abrir en Excel
    wb_dest.calculation.calcMode = 'auto'

    print(f"\n{'='*65}")
    try:
        wb_dest.save(ARCHIVO_PRINCIPAL)
        print(f"[OK] Archivo guardado: {ARCHIVO_PRINCIPAL}")
        # Registrar ejecución exitosa en la DB
        todos_log = (log_gastos or [])
        _db.log_ejecucion(mes_nombre, exito=True, valores=todos_log)
        print(f"[DB] Ejecución registrada para {mes_nombre}")
    except PermissionError:
        print(f"[ERROR] Cierre el archivo en Excel y vuelva a ejecutar.")
        _db.log_ejecucion(mes_nombre, exito=False, valores=[],
                          notas="PermissionError al guardar el archivo")


# ---------------------------------------------------------------------------
if __name__ == '__main__':
    if len(sys.argv) > 1:
        arg = sys.argv[1]
        carpeta = arg if os.path.isabs(arg) else os.path.join(BASE, arg)
    else:
        carpeta = os.path.join(BASE, 'FEBRERO-DL')

    procesar_mes(carpeta)
